from __future__ import annotations

import os
import queue
import sys
import threading
import time
import ctypes
from datetime import date as dt_date, datetime, timedelta
import math
from pathlib import Path
import unicodedata

import customtkinter as ctk
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from services.exceptions import (
    ExceptionConfigError,
    append_manual_exceptions_to_file,
    ensure_exceptions_file,
)
from services.processor import ProcessorService
from services.parser import load_hikvision_excel

try:
    from tkcalendar import Calendar
except Exception:  # pragma: no cover - optional dependency at runtime
    Calendar = None

DEFAULT_ABSENCE_TYPES = [
    "DOMINGO",
    "AUSENCIA",
    "VACACIONES",
    "ESTUDIAR",
    "CAPACITACIONES",
    "SUSPENCION",
    "NO TRABAJADO",
    "LICENCIA",
    "FERIADO",
    "ACCIDENTE DE TRABAJO",
]
DEFAULT_ABSENCE_STATES = ["PENDIENTE", "APROBADO", "RECHAZADO"]


def get_app_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent.parent

COLOR_BG = ("#EEF3F8", "#0A1220")
COLOR_PANEL = ("#FFFFFF", "#111827")
COLOR_PANEL_SOFT = ("#F8FAFC", "#0F172A")
COLOR_BORDER = ("#D7E1ED", "#243244")
COLOR_TEXT = ("#0F172A", "#E5E7EB")
COLOR_TEXT_MUTED = ("#475569", "#93A4B8")
COLOR_PRIMARY = ("#0E7490", "#22D3EE")
COLOR_PRIMARY_HOVER = ("#155E75", "#06B6D4")
COLOR_SECONDARY_HOVER = ("#E2E8F0", "#1E293B")
COLOR_TAB_ACTIVE = ("#E0F2FE", "#10243E")
COLOR_TAB_TEXT_ACTIVE = ("#0C4A6E", "#67E8F9")
COLOR_STATUS_READY_BG = ("#DCFCE7", "#133124")
COLOR_STATUS_READY_TEXT = ("#166534", "#86EFAC")
COLOR_STATUS_WORKING_BG = ("#FEF3C7", "#3F3208")
COLOR_STATUS_WORKING_TEXT = ("#92400E", "#FDE68A")
COLOR_STATUS_WORKING_BG_ALT = ("#FDE68A", "#4A3A0A")
COLOR_STATUS_ERROR_BG = ("#FEE2E2", "#3D1111")
COLOR_STATUS_ERROR_TEXT = ("#991B1B", "#FCA5A5")
COLOR_KPI_NEUTRAL = ("#D7EEF7", "#12384A")
COLOR_KPI_POSITIVE = ("#DCFCE7", "#163823")
COLOR_KPI_ALERT = ("#FEE2E2", "#3A1616")
COLOR_KPI_WARNING = ("#FEF3C7", "#3B3312")
COLOR_CHIP_NEUTRAL_BG = ("#E2E8F0", "#1F2937")
COLOR_CHIP_NEUTRAL_TEXT = ("#334155", "#CBD5E1")
COLOR_CHIP_OK_BG = ("#DCFCE7", "#173124")
COLOR_CHIP_OK_TEXT = ("#166534", "#86EFAC")
COLOR_CHIP_WARN_BG = ("#FEF3C7", "#3F3208")
COLOR_CHIP_WARN_TEXT = ("#92400E", "#FDE68A")
COLOR_CHIP_ERROR_BG = ("#FEE2E2", "#3D1111")
COLOR_CHIP_ERROR_TEXT = ("#991B1B", "#FCA5A5")
COLOR_CHIP_WORKING_BG = ("#E0F2FE", "#0B2740")
COLOR_CHIP_WORKING_TEXT = ("#0C4A6E", "#67E8F9")
COLOR_LOG_INFO = ("#0F172A", "#E5E7EB")
COLOR_LOG_SUCCESS = ("#166534", "#86EFAC")
COLOR_LOG_WARN = ("#92400E", "#FDE68A")
COLOR_LOG_ERROR = ("#B91C1C", "#FCA5A5")


class HikvisionApp(ctk.CTk):
    def __init__(self) -> None:
        super().__init__()
        self.app_base_dir = get_app_base_dir()
        if getattr(sys, "frozen", False):
            os.chdir(self.app_base_dir)

        self.title("Hikvision Hours Processor")
        self.geometry("1220x760")
        self.minsize(1080, 640)
        self._apply_app_icon()

        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("dark-blue")
        self.configure(fg_color=COLOR_BG)

        self.font_title = ctk.CTkFont(family="Bahnschrift", size=26, weight="bold")
        self.font_section = ctk.CTkFont(family="Bahnschrift", size=20, weight="bold")
        self.font_subtitle = ctk.CTkFont(family="Bahnschrift", size=13)
        self.font_body = ctk.CTkFont(family="Bahnschrift", size=12)
        self.font_button = ctk.CTkFont(family="Bahnschrift", size=13, weight="bold")
        self.font_mono = ctk.CTkFont(family="Consolas", size=12)

        self.processor = ProcessorService()
        self.selected_file: Path | None = None
        self.exceptions_file: Path | None = None
        self.output_file: Path | None = None

        self._worker_thread: threading.Thread | None = None
        self._progress_queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self._started_at: float | None = None

        self.current_page = "report"

        self.log_box: ctk.CTkTextbox | None = None
        self.exceptions_file_entry: ctk.CTkEntry | None = None
        self.exceptions_manual_box: ctk.CTkTextbox | None = None
        self.quick_employee_menu: ctk.CTkOptionMenu | None = None
        self.quick_date_entry: ctk.CTkEntry | None = None
        self.quick_date_end_entry: ctk.CTkEntry | None = None
        self.quick_date_button: ctk.CTkButton | None = None
        self.quick_type_menu: ctk.CTkOptionMenu | None = None
        self.quick_detail_entry: ctk.CTkEntry | None = None
        self.quick_add_button: ctk.CTkButton | None = None
        self.quick_remove_last_button: ctk.CTkButton | None = None
        self.quick_clear_manual_button: ctk.CTkButton | None = None
        self.exceptions_summary_label: ctk.CTkLabel | None = None
        self.open_date_button_report: ctk.CTkButton | None = None
        self.open_date_button_exceptions: ctk.CTkButton | None = None
        self.download_study_button: ctk.CTkButton | None = None
        self.context_file_value: ctk.CTkLabel | None = None
        self.context_exceptions_value: ctk.CTkLabel | None = None
        self.context_manual_value: ctk.CTkLabel | None = None
        self.context_output_value: ctk.CTkLabel | None = None
        self.kpi_employees_value: ctk.CTkLabel | None = None
        self.kpi_days_value: ctk.CTkLabel | None = None
        self.kpi_late_value: ctk.CTkLabel | None = None
        self.kpi_absent_value: ctk.CTkLabel | None = None
        self.kpi_overtime_value: ctk.CTkLabel | None = None
        self._logo_image: tk.PhotoImage | None = None
        self._logo_badge: ctk.CTkFrame | None = None
        self._logo_label: ctk.CTkLabel | None = None
        self._kpi_cards: dict[str, dict[str, object]] = {}
        self.report_hero_title: ctk.CTkLabel | None = None
        self.report_hero_subtitle: ctk.CTkLabel | None = None
        self.health_chip_file: ctk.CTkLabel | None = None
        self.health_chip_exceptions: ctk.CTkLabel | None = None
        self.health_chip_run: ctk.CTkLabel | None = None
        self.exceptions_chip_file: ctk.CTkLabel | None = None
        self.exceptions_chip_manual: ctk.CTkLabel | None = None
        self.exceptions_chip_save: ctk.CTkLabel | None = None
        self._log_text_widget: tk.Text | None = None
        self._page_frames: dict[str, ctk.CTkFrame] = {}
        self._page_transition_after: str | None = None
        self._page_transition_running: bool = False
        self._pending_page: str | None = None
        self._employee_selector_values: list[str] = ["Todos"]
        self._employee_label_to_id: dict[str, str] = {"Todos": ""}
        self._absence_type_values: list[str] = list(DEFAULT_ABSENCE_TYPES)
        self._status_pulse_job: str | None = None
        self._status_pulse_on: bool = False

        self._build_ui()
        self._initialize_default_exceptions_file()
        self._refresh_absence_type_options()
        self._sync_absence_types_dropdown_in_template()
        self._refresh_employee_options()

    def _apply_app_icon(self) -> None:
        base_dir = get_app_base_dir()
        ico_path = base_dir / "app.ico"
        png_path = base_dir / "logo.png"

        if sys.platform.startswith("win"):
            try:
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                    "Porta.HikvisionHoursProcessor"
                )
            except Exception:
                pass

        if ico_path.exists():
            try:
                self.iconbitmap(str(ico_path))
            except Exception:
                pass

        if png_path.exists():
            try:
                icon_image = tk.PhotoImage(file=str(png_path))
                self.iconphoto(True, icon_image)
                self._window_icon_image = icon_image
            except Exception:
                pass

    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.navbar = ctk.CTkFrame(self, corner_radius=0, height=66, fg_color=("#DCE8F6", "#0E1A2B"))
        self.navbar.grid(row=0, column=0, sticky="ew")
        self.navbar.grid_columnconfigure(0, weight=0)
        self.navbar.grid_columnconfigure(1, weight=0)
        self.navbar.grid_columnconfigure(2, weight=0)
        self.navbar.grid_columnconfigure(3, weight=1)
        self.navbar.grid_columnconfigure(4, weight=0)
        self.navbar.grid_rowconfigure(0, weight=0)
        self.navbar.grid_rowconfigure(1, weight=0, minsize=3)

        self.nav_title = ctk.CTkLabel(
            self.navbar,
            text="Hikvision Hours",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLOR_TEXT,
        )
        self.nav_title.grid(row=0, column=0, padx=(14, 18), pady=(10, 8), sticky="w")

        self.report_tab_button = ctk.CTkButton(
            self.navbar,
            text="Reporte",
            width=110,
            height=32,
            corner_radius=6,
            fg_color="transparent",
            hover_color=COLOR_SECONDARY_HOVER,
            text_color=COLOR_TEXT_MUTED,
            font=self.font_button,
            command=lambda: self._show_page("report"),
        )
        self.report_tab_button.grid(row=0, column=1, padx=(0, 6), pady=(12, 8), sticky="w")

        self.exceptions_tab_button = ctk.CTkButton(
            self.navbar,
            text="Excepciones",
            width=120,
            height=32,
            corner_radius=6,
            fg_color="transparent",
            hover_color=COLOR_SECONDARY_HOVER,
            text_color=COLOR_TEXT_MUTED,
            font=self.font_button,
            command=lambda: self._show_page("exceptions"),
        )
        # self.exceptions_tab_button.grid(row=0, column=2, padx=(0, 6), pady=(12, 8), sticky="w")  # hidden

        self.employees_tab_button = ctk.CTkButton(
            self.navbar,
            text="Empleados",
            width=110,
            height=32,
            corner_radius=6,
            fg_color="transparent",
            hover_color=COLOR_SECONDARY_HOVER,
            text_color=COLOR_TEXT_MUTED,
            font=self.font_button,
            command=lambda: self._show_page("employees"),
        )
        self.employees_tab_button.grid(row=0, column=3, padx=(0, 10), pady=(12, 8), sticky="w")

        self.report_indicator = ctk.CTkFrame(
            self.navbar,
            height=3,
            width=110,
            corner_radius=0,
            fg_color=COLOR_PRIMARY,
        )
        self.report_indicator.grid(row=1, column=1, padx=(0, 6), sticky="sw")

        self.exceptions_indicator = ctk.CTkFrame(
            self.navbar,
            height=3,
            width=120,
            corner_radius=0,
            fg_color=COLOR_PRIMARY,
        )
        # self.exceptions_indicator.grid(row=1, column=2, padx=(0, 6), sticky="sw")  # hidden

        self.employees_indicator = ctk.CTkFrame(
            self.navbar,
            height=3,
            width=110,
            corner_radius=0,
            fg_color=COLOR_PRIMARY,
        )
        self.employees_indicator.grid(row=1, column=3, padx=(0, 10), sticky="sw")

        self._mount_logo()

        self.content = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.content.grid(row=1, column=0, sticky="nsew")
        self.content.grid_columnconfigure(0, weight=1)
        self.content.grid_rowconfigure(0, weight=1)

        self.report_page = self._build_report_page(self.content)
        self.exceptions_page = self._build_exceptions_page(self.content)
        self.employees_page = self._build_employees_page(self.content)
        self._page_frames = {
            "report": self.report_page,
            # "exceptions": self.exceptions_page,  # hidden
            "employees": self.employees_page,
        }

        self.report_page.place(x=0, y=0, relwidth=1, relheight=1)
        # self.exceptions_page.place(x=0, y=0, relwidth=1, relheight=1)  # hidden
        self.employees_page.place(x=0, y=0, relwidth=1, relheight=1)
        self.report_page.lift()
        self.exceptions_page.lower()
        self.employees_page.lower()

        self._reset_kpis()
        self._refresh_exceptions_summary()
        self._refresh_context_panel()
        self._show_page("report", animate=False)

    def _style_primary_button(self, button: ctk.CTkButton, *, height: int | None = None) -> None:
        kwargs: dict[str, object] = {
            "fg_color": COLOR_PRIMARY,
            "hover_color": COLOR_PRIMARY_HOVER,
            "text_color": ("#FFFFFF", "#06202A"),
            "font": self.font_button,
            "corner_radius": 10,
        }
        if height is not None:
            kwargs["height"] = height
        button.configure(**kwargs)

    def _style_secondary_button(self, button: ctk.CTkButton, *, height: int | None = None) -> None:
        kwargs: dict[str, object] = {
            "fg_color": "transparent",
            "hover_color": COLOR_SECONDARY_HOVER,
            "border_width": 1,
            "border_color": COLOR_BORDER,
            "text_color": COLOR_TEXT,
            "font": self.font_body,
            "corner_radius": 10,
        }
        if height is not None:
            kwargs["height"] = height
        button.configure(**kwargs)

    def _set_context_value(self, label: ctk.CTkLabel | None, value: str) -> None:
        if label is not None:
            label.configure(text=value)

    def _theme_color(self, pair: tuple[str, str] | str) -> str:
        if isinstance(pair, str):
            return pair
        mode = ctk.get_appearance_mode().lower()
        return pair[1] if mode == "dark" else pair[0]

    def _set_chip_state(self, chip: ctk.CTkLabel | None, text: str, level: str) -> None:
        if chip is None:
            return
        palette = {
            "neutral": {"fg": COLOR_CHIP_NEUTRAL_BG, "text": COLOR_CHIP_NEUTRAL_TEXT},
            "ok": {"fg": COLOR_CHIP_OK_BG, "text": COLOR_CHIP_OK_TEXT},
            "warn": {"fg": COLOR_CHIP_WARN_BG, "text": COLOR_CHIP_WARN_TEXT},
            "error": {"fg": COLOR_CHIP_ERROR_BG, "text": COLOR_CHIP_ERROR_TEXT},
            "working": {"fg": COLOR_CHIP_WORKING_BG, "text": COLOR_CHIP_WORKING_TEXT},
        }
        colors = palette.get(level, palette["neutral"])
        chip.configure(text=text, fg_color=colors["fg"], text_color=colors["text"])

    def _refresh_health_chips(self) -> None:
        file_loaded = self.selected_file is not None
        exceptions_loaded = self.exceptions_file is not None
        is_running = bool(self._worker_thread and self._worker_thread.is_alive())
        has_output = self.output_file is not None

        self._set_chip_state(
            self.health_chip_file,
            "Archivo: cargado" if file_loaded else "Archivo: pendiente",
            "ok" if file_loaded else "warn",
        )
        self._set_chip_state(
            self.health_chip_exceptions,
            "Excepciones: listas" if exceptions_loaded else "Excepciones: pendientes",
            "ok" if exceptions_loaded else "warn",
        )
        if is_running:
            run_text = "Ejecucion: en proceso"
            run_level = "working"
        elif has_output:
            run_text = "Ejecucion: completada"
            run_level = "ok"
        else:
            run_text = "Ejecucion: sin iniciar"
            run_level = "neutral"
        self._set_chip_state(self.health_chip_run, run_text, run_level)
        self._refresh_exceptions_hero_chips()

    def _refresh_exceptions_hero_chips(self) -> None:
        file_text = self.exceptions_file.name if self.exceptions_file else "sin archivo"
        self._set_chip_state(
            self.exceptions_chip_file,
            f"Archivo: {file_text}",
            "ok" if self.exceptions_file is not None else "warn",
        )

        manual_text = self.exceptions_manual_box.get("1.0", "end").strip() if self.exceptions_manual_box else ""
        manual_count = self._count_manual_exceptions(manual_text)
        manual_level = "neutral" if manual_count == 0 else "working" if manual_count <= 5 else "warn"
        self._set_chip_state(
            self.exceptions_chip_manual,
            f"Pendiente manual: {manual_count} lineas",
            manual_level,
        )

        if self._worker_thread and self._worker_thread.is_alive():
            save_text = "Guardado: bloqueado (procesando)"
            save_level = "warn"
        else:
            save_text = "Guardado: disponible"
            save_level = "ok"
        self._set_chip_state(self.exceptions_chip_save, save_text, save_level)

    def _refresh_page_titles(self, page: str | None = None) -> None:
        active_page = page or self.current_page
        titles = {
            "report": "Hikvision Hours · Reporte",
            "exceptions": "Hikvision Hours · Excepciones",
            "employees": "Hikvision Hours · Empleados",
        }
        self.nav_title.configure(text=titles.get(active_page, "Hikvision Hours"))

    def _apply_tab_visual_state(self, page: str) -> None:
        active_tab_fg = COLOR_TAB_ACTIVE
        inactive_tab_fg = "transparent"
        active_text = COLOR_TAB_TEXT_ACTIVE
        inactive_text = COLOR_TEXT_MUTED
        active_line = COLOR_PRIMARY
        hidden_line = ("#DCE8F6", "#0E1A2B")

        all_buttons = [
            (self.report_tab_button, self.report_indicator, "report"),
            # (self.exceptions_tab_button, self.exceptions_indicator, "exceptions"),  # hidden
            (self.employees_tab_button, self.employees_indicator, "employees"),
        ]
        for btn, ind, name in all_buttons:
            if name == page:
                btn.configure(fg_color=active_tab_fg, text_color=active_text)
                ind.configure(fg_color=active_line)
            else:
                btn.configure(fg_color=inactive_tab_fg, text_color=inactive_text)
                ind.configure(fg_color=hidden_line)

    def _finish_page_transition(self, target_page: str, outgoing_page: str) -> None:
        target_frame = self._page_frames[target_page]
        outgoing_frame = self._page_frames[outgoing_page]
        target_frame.place_configure(x=0, y=0, relwidth=1, relheight=1)
        outgoing_frame.place_configure(x=0, y=0, relwidth=1, relheight=1)
        target_frame.lift()
        outgoing_frame.lower()
        self.current_page = target_page
        self._page_transition_running = False
        self._page_transition_after = None
        self._refresh_page_titles(target_page)

        if self._pending_page and self._pending_page != self.current_page:
            pending = self._pending_page
            self._pending_page = None
            self._show_page(pending, animate=True)
        else:
            self._pending_page = None

    def _ease_in_out_cubic(self, t: float) -> float:
        if t <= 0:
            return 0.0
        if t >= 1:
            return 1.0
        if t < 0.5:
            return 4 * t * t * t
        return 1 - pow(-2 * t + 2, 3) / 2

    def _ease_out_cubic(self, t: float) -> float:
        if t <= 0:
            return 0.0
        if t >= 1:
            return 1.0
        return 1 - pow(1 - t, 3)

    def _animate_page_switch(self, target_page: str) -> None:
        if target_page not in self._page_frames:
            return

        if self._page_transition_running:
            self._pending_page = target_page
            return

        outgoing_page = self.current_page
        if outgoing_page == target_page:
            return
        if outgoing_page not in self._page_frames:
            self.current_page = target_page
            self._page_frames[target_page].lift()
            self._refresh_page_titles(target_page)
            return

        self._page_transition_running = True
        self._pending_page = None
        self._apply_tab_visual_state(target_page)
        self._refresh_page_titles(target_page)

        outgoing_frame = self._page_frames[outgoing_page]
        target_frame = self._page_frames[target_page]

        self.update_idletasks()
        width = max(1, self.content.winfo_width())
        direction = 1 if target_page == "exceptions" else -1
        steps = 24
        delay_ms = 11
        settle_steps = 6
        settle_px = 12

        start_out = 0
        end_out = -direction * max(140, int(width * 0.20))
        start_in = direction * max(220, int(width * 0.38))
        end_in = 0

        outgoing_frame.place_configure(x=start_out, y=0, relwidth=1, relheight=1)
        target_frame.place_configure(x=start_in, y=0, relwidth=1, relheight=1)
        target_frame.lift()

        if self._page_transition_after is not None:
            self.after_cancel(self._page_transition_after)
            self._page_transition_after = None

        def _step(index: int) -> None:
            t = index / steps
            eased = self._ease_in_out_cubic(t)
            out_x = round(start_out + (end_out - start_out) * eased)
            in_x = round(start_in + (end_in - start_in) * eased)

            outgoing_frame.place_configure(x=out_x, y=0, relwidth=1, relheight=1)
            target_frame.place_configure(x=in_x, y=0, relwidth=1, relheight=1)

            if index < steps:
                self._page_transition_after = self.after(delay_ms, lambda: _step(index + 1))
                return

            overshoot = -direction * settle_px
            target_frame.place_configure(x=overshoot, y=0, relwidth=1, relheight=1)

            def _settle(step_idx: int) -> None:
                tt = step_idx / settle_steps
                settle_eased = self._ease_out_cubic(tt)
                x = round(overshoot * (1 - settle_eased))
                target_frame.place_configure(x=x, y=0, relwidth=1, relheight=1)
                if step_idx < settle_steps:
                    self._page_transition_after = self.after(delay_ms, lambda: _settle(step_idx + 1))
                else:
                    self._finish_page_transition(target_page, outgoing_page)

            _settle(0)

        _step(0)

    def _minutes_to_hhmm(self, total_minutes: int) -> str:
        hours, minutes = divmod(max(0, int(total_minutes)), 60)
        return f"{hours:02d}:{minutes:02d}"

    def _set_kpi_value(self, label: ctk.CTkLabel | None, value: str) -> None:
        if label is not None:
            label.configure(text=value)

    def _to_non_negative_int(self, value: object) -> int:
        try:
            return max(0, int(str(value).strip()))
        except (TypeError, ValueError):
            return 0

    def _hhmm_to_minutes(self, value: str) -> int:
        if ":" not in value:
            return 0
        try:
            hours, minutes = value.split(":", 1)
            return max(0, int(hours) * 60 + int(minutes))
        except ValueError:
            return 0

    def _set_kpi_card_style(self, key: str, variant: str) -> None:
        card_data = self._kpi_cards.get(key)
        if not card_data:
            return

        palette = {
            "neutral": COLOR_KPI_NEUTRAL,
            "positive": COLOR_KPI_POSITIVE,
            "warning": COLOR_KPI_WARNING,
            "alert": COLOR_KPI_ALERT,
        }
        color_pair = palette.get(variant, COLOR_KPI_NEUTRAL)
        accent = card_data.get("accent")
        if isinstance(accent, ctk.CTkFrame):
            accent.configure(fg_color=color_pair)

    def _refresh_kpi_visual_state(self) -> None:
        employees_count = self._to_non_negative_int(self.kpi_employees_value.cget("text")) if self.kpi_employees_value else 0
        days_count = self._to_non_negative_int(self.kpi_days_value.cget("text")) if self.kpi_days_value else 0
        late_count = self._to_non_negative_int(self.kpi_late_value.cget("text")) if self.kpi_late_value else 0
        absent_count = self._to_non_negative_int(self.kpi_absent_value.cget("text")) if self.kpi_absent_value else 0
        overtime_minutes = self._hhmm_to_minutes(self.kpi_overtime_value.cget("text")) if self.kpi_overtime_value else 0

        self._set_kpi_card_style("employees", "positive" if employees_count > 0 else "neutral")
        self._set_kpi_card_style("days", "positive" if days_count > 0 else "neutral")
        self._set_kpi_card_style("late", "warning" if late_count > 0 else "positive")
        self._set_kpi_card_style("absent", "alert" if absent_count > 0 else "positive")
        self._set_kpi_card_style("overtime", "neutral" if overtime_minutes <= 0 else "warning")

    def _marker_is_positive(self, value: object) -> bool:
        text = str(value).strip()
        if text in {"", "-", "0", "0.0", "0:0", "0:00", "nan", "NaN", "None"}:
            return False
        if ":" in text:
            parts = [p.strip() for p in text.split(":") if p.strip() != ""]
            if not parts:
                return False
            try:
                return any(int(p) > 0 for p in parts)
            except ValueError:
                return True
        try:
            return float(text.replace(",", ".")) > 0
        except ValueError:
            return True

    def _log_level_for_message(self, message: str) -> str:
        lowered = message.strip().lower()
        if not lowered:
            return "info"
        if any(token in lowered for token in ["error", "fallo", "invalido", "no se pudo"]):
            return "error"
        if any(token in lowered for token in ["warning", "atencion", "ausente", "tarde"]):
            return "warn"
        if any(token in lowered for token in ["completado", "generado", "guardada", "listo"]):
            return "success"
        return "info"

    def _reset_kpis(self) -> None:
        self._set_kpi_value(self.kpi_employees_value, "-")
        self._set_kpi_value(self.kpi_days_value, "-")
        self._set_kpi_value(self.kpi_late_value, "-")
        self._set_kpi_value(self.kpi_absent_value, "-")
        self._set_kpi_value(self.kpi_overtime_value, "-")
        self._refresh_kpi_visual_state()

    def _refresh_kpis_from_daily_df(self, daily_df: pd.DataFrame) -> None:
        if daily_df.empty:
            self._reset_kpis()
            return

        employees = int(daily_df["ID de persona"].astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if "ID de persona" in daily_df.columns else 0
        days = int(len(daily_df))
        status_series = daily_df["Estado"].astype(str).str.strip().str.lower() if "Estado" in daily_df.columns else pd.Series(dtype=str)
        late_count = int(status_series.isin(["tarde", "tardanza"]).sum())
        absent_count = int(status_series.isin(["ausente", "ausencia"]).sum())
        overtime_minutes = (
            pd.to_numeric(daily_df["Minutos extra"], errors="coerce").fillna(0).astype(int).sum()
            if "Minutos extra" in daily_df.columns
            else 0
        )

        self._set_kpi_value(self.kpi_employees_value, str(employees))
        self._set_kpi_value(self.kpi_days_value, str(days))
        self._set_kpi_value(self.kpi_late_value, str(late_count))
        self._set_kpi_value(self.kpi_absent_value, str(absent_count))
        self._set_kpi_value(self.kpi_overtime_value, self._minutes_to_hhmm(int(overtime_minutes)))
        self._refresh_kpi_visual_state()

    def _refresh_kpis_from_source(self, source_path: Path) -> None:
        try:
            df = load_hikvision_excel(source_path)
        except Exception:
            self._reset_kpis()
            return

        if df.empty:
            self._reset_kpis()
            return

        employees = int(df["employee_id"].astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if "employee_id" in df.columns else 0
        days = int(pd.to_datetime(df["work_date_raw"], errors="coerce").dropna().dt.date.nunique()) if "work_date_raw" in df.columns else 0
        late_count = int(df["late_raw"].map(self._marker_is_positive).sum()) if "late_raw" in df.columns else 0
        absent_count = int(df["absent_raw"].map(self._marker_is_positive).sum()) if "absent_raw" in df.columns else 0

        self._set_kpi_value(self.kpi_employees_value, str(employees))
        self._set_kpi_value(self.kpi_days_value, str(days))
        self._set_kpi_value(self.kpi_late_value, str(late_count))
        self._set_kpi_value(self.kpi_absent_value, str(absent_count))
        self._set_kpi_value(self.kpi_overtime_value, "--:--")
        self._refresh_kpi_visual_state()

    def _refresh_kpis_from_output(self, output_path: Path) -> None:
        try:
            daily_df = pd.read_excel(output_path, sheet_name="Diario")
        except Exception:
            return
        self._refresh_kpis_from_daily_df(daily_df)

    def _refresh_context_panel(self) -> None:
        file_value = self.selected_file.name if self.selected_file else "Sin archivo"
        exc_value = self.exceptions_file.name if self.exceptions_file else "Sin archivo"
        manual_text = self.exceptions_manual_box.get("1.0", "end").strip() if self.exceptions_manual_box else ""
        manual_value = f"{self._count_manual_exceptions(manual_text)} lineas"
        output_value = self.output_file.name if self.output_file else "No generado"

        self._set_context_value(self.context_file_value, file_value)
        self._set_context_value(self.context_exceptions_value, exc_value)
        self._set_context_value(self.context_manual_value, manual_value)
        self._set_context_value(self.context_output_value, output_value)
        self._refresh_health_chips()

    def _start_status_pulse(self) -> None:
        if self._status_pulse_job is not None:
            return

        def _pulse() -> None:
            self._status_pulse_on = not self._status_pulse_on
            self.status_badge.configure(
                fg_color=COLOR_STATUS_WORKING_BG_ALT if self._status_pulse_on else COLOR_STATUS_WORKING_BG
            )
            self._status_pulse_job = self.after(500, _pulse)

        self._status_pulse_on = False
        _pulse()

    def _stop_status_pulse(self) -> None:
        if self._status_pulse_job is not None:
            self.after_cancel(self._status_pulse_job)
            self._status_pulse_job = None
        self._status_pulse_on = False

    def _mount_logo(self) -> None:
        logo_path = get_app_base_dir() / "logo.png"
        if not logo_path.exists():
            return

        try:
            raw_logo = tk.PhotoImage(file=str(logo_path))
        except Exception:
            return

        max_width = 150
        max_height = 30
        ratio = max(raw_logo.width() / max_width, raw_logo.height() / max_height, 1)
        scale = max(1, math.ceil(ratio))

        self._logo_image = raw_logo.subsample(scale, scale) if scale > 1 else raw_logo
        self._logo_badge = ctk.CTkFrame(
            self.navbar,
            corner_radius=12,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=("#F8FBFF", "#111827"),
            width=190,
            height=40,
        )
        self._logo_badge.grid(row=0, column=4, padx=(8, 14), pady=(10, 8), sticky="e")
        self._logo_badge.grid_propagate(False)

        self._logo_label = ctk.CTkLabel(
            self._logo_badge,
            text="",
            image=self._logo_image,
            fg_color="transparent",
        )
        self._logo_label.place(relx=0.5, rely=0.5, anchor="center")

    def _build_report_page(self, parent: ctk.CTkFrame) -> ctk.CTkFrame:
        page = ctk.CTkFrame(parent, corner_radius=0, fg_color="transparent")
        page.grid_columnconfigure(0, weight=0)
        page.grid_columnconfigure(1, weight=1)
        page.grid_rowconfigure(0, weight=1)

        sidebar = ctk.CTkFrame(
            page,
            width=370,
            corner_radius=18,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL,
        )
        sidebar.grid(row=0, column=0, sticky="nsw", padx=(14, 10), pady=12)
        sidebar.grid_propagate(False)
        sidebar.grid_columnconfigure(0, weight=1)

        title = ctk.CTkLabel(
            sidebar,
            text="Procesador Hikvision",
            font=self.font_title,
            text_color=COLOR_TEXT,
        )
        title.grid(row=0, column=0, padx=20, pady=(22, 8), sticky="w")

        subtitle = ctk.CTkLabel(
            sidebar,
            text="Importa fichadas y genera reportes\ndiarios/mensuales.",
            justify="left",
            text_color=COLOR_TEXT_MUTED,
            font=self.font_subtitle,
        )
        subtitle.grid(row=1, column=0, padx=20, pady=(0, 16), sticky="w")

        self.status_badge = ctk.CTkLabel(
            sidebar,
            text="Estado: listo",
            corner_radius=8,
            fg_color=COLOR_STATUS_READY_BG,
            text_color=COLOR_STATUS_READY_TEXT,
            padx=10,
            pady=6,
            font=self.font_body,
        )
        self.status_badge.grid(row=2, column=0, padx=20, pady=(0, 14), sticky="w")

        select_label = ctk.CTkLabel(
            sidebar, text="Archivo de entrada", font=self.font_button, text_color=COLOR_TEXT
        )
        select_label.grid(row=3, column=0, padx=20, pady=(0, 8), sticky="w")

        self.file_entry = ctk.CTkEntry(sidebar, placeholder_text="Ningun archivo seleccionado")
        self.file_entry.grid(row=4, column=0, padx=20, pady=(0, 8), sticky="ew")
        self.file_entry.configure(border_color=COLOR_BORDER, fg_color=COLOR_PANEL_SOFT, text_color=COLOR_TEXT)
        self.file_entry.configure(state="disabled")

        self.file_meta = ctk.CTkLabel(
            sidebar,
            text="",
            justify="left",
            text_color=COLOR_TEXT_MUTED,
            font=self.font_body,
        )
        self.file_meta.grid(row=5, column=0, padx=20, pady=(0, 12), sticky="w")

        self.select_button = ctk.CTkButton(
            sidebar,
            text="Seleccionar archivo",
            height=38,
            command=self.select_file,
        )
        self._style_primary_button(self.select_button, height=38)
        self.select_button.grid(row=6, column=0, padx=20, pady=(0, 10), sticky="ew")

        self.open_date_button_report = ctk.CTkButton(
            sidebar,
            text="Abrir date.xlsx",
            height=34,
            command=self.open_date_template_file,
        )
        self._style_secondary_button(self.open_date_button_report, height=34)
        self.open_date_button_report.grid(row=7, column=0, padx=20, pady=(0, 8), sticky="ew")

        self.exceptions_summary_label = ctk.CTkLabel(
            sidebar,
            text="Excepciones: sin configurar",
            justify="left",
            text_color=COLOR_TEXT_MUTED,
            font=self.font_body,
        )
        self.exceptions_summary_label.grid(row=8, column=0, padx=20, pady=(0, 8), sticky="w")

        self.process_button = ctk.CTkButton(
            sidebar,
            text="Procesar reporte",
            height=40,
            command=self.process_file,
            state="disabled",
        )
        self._style_primary_button(self.process_button, height=40)
        self.process_button.grid(row=9, column=0, padx=20, pady=(2, 10), sticky="ew")

        self.open_button = ctk.CTkButton(
            sidebar,
            text="Abrir reporte generado",
            height=36,
            command=self.open_output_file,
            state="disabled",
        )
        self._style_primary_button(self.open_button, height=36)
        self.open_button.grid(row=10, column=0, padx=20, pady=(0, 8), sticky="ew")

        self.download_study_button = ctk.CTkButton(
            sidebar,
            text="Descargar reporte estudio",
            height=36,
            command=self.download_study_report,
            state="disabled",
        )
        self._style_primary_button(self.download_study_button, height=36)
        self.download_study_button.grid(row=11, column=0, padx=20, pady=(0, 8), sticky="ew")

        self.clear_button = ctk.CTkButton(
            sidebar,
            text="Limpiar bitacora",
            height=36,
            command=self.clear_log,
        )
        self._style_secondary_button(self.clear_button, height=36)
        self.clear_button.grid(row=12, column=0, padx=20, pady=(0, 14), sticky="ew")

        self.progress = ctk.CTkProgressBar(sidebar, mode="determinate")
        self.progress.configure(progress_color=COLOR_PRIMARY, fg_color=("#DCEAF6", "#1F2937"))
        self.progress.grid(row=13, column=0, padx=20, pady=(0, 8), sticky="ew")
        self.progress.set(0)

        self.progress_text = ctk.CTkLabel(sidebar, text="Progreso: 0%", text_color=COLOR_TEXT_MUTED, font=self.font_body)
        self.progress_text.grid(row=14, column=0, padx=20, pady=(0, 4), sticky="w")

        self.progress_step = ctk.CTkLabel(
            sidebar,
            text="Paso: esperando archivo",
            text_color=COLOR_TEXT_MUTED,
            font=self.font_body,
        )
        self.progress_step.grid(row=15, column=0, padx=20, pady=(0, 4), sticky="w")

        self.elapsed_text = ctk.CTkLabel(sidebar, text="Tiempo: 00:00", text_color=COLOR_TEXT_MUTED, font=self.font_body)
        self.elapsed_text.grid(row=16, column=0, padx=20, pady=(0, 10), sticky="w")

        info = ctk.CTkLabel(
            sidebar,
            text="Entrada: .xls/.xlsx",
            text_color=COLOR_TEXT_MUTED,
            justify="left",
            font=self.font_body,
        )
        info.grid(row=17, column=0, padx=20, pady=(0, 12), sticky="w")

        main = ctk.CTkFrame(
            page,
            corner_radius=18,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL,
        )
        main.grid(row=0, column=1, sticky="nsew", padx=(8, 14), pady=12)
        main.grid_columnconfigure(0, weight=1)
        main.grid_rowconfigure(4, weight=1)

        hero = ctk.CTkFrame(
            main,
            corner_radius=12,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL_SOFT,
        )
        hero.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        hero.grid_columnconfigure(0, weight=1)
        hero.grid_rowconfigure(2, weight=0)

        hero_accent = ctk.CTkFrame(hero, corner_radius=8, height=6, fg_color=COLOR_PRIMARY)
        hero_accent.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 0))

        self.report_hero_title = ctk.CTkLabel(
            hero,
            text="Centro de Operaciones",
            font=self.font_section,
            text_color=COLOR_TEXT,
        )
        self.report_hero_title.grid(row=1, column=0, sticky="w", padx=12, pady=(8, 2))

        self.report_hero_subtitle = ctk.CTkLabel(
            hero,
            text="Bitacora en vivo del procesamiento, validaciones y reporte final.",
            text_color=COLOR_TEXT_MUTED,
            font=self.font_subtitle,
        )
        self.report_hero_subtitle.grid(row=2, column=0, sticky="w", padx=12, pady=(0, 8))

        health_row = ctk.CTkFrame(hero, fg_color="transparent")
        health_row.grid(row=3, column=0, sticky="ew", padx=12, pady=(0, 10))
        health_row.grid_columnconfigure((0, 1, 2), weight=1)

        def _build_chip(parent: ctk.CTkFrame, text: str) -> ctk.CTkLabel:
            chip = ctk.CTkLabel(
                parent,
                text=text,
                corner_radius=8,
                fg_color=COLOR_CHIP_NEUTRAL_BG,
                text_color=COLOR_CHIP_NEUTRAL_TEXT,
                padx=8,
                pady=5,
                font=self.font_body,
            )
            return chip

        self.health_chip_file = _build_chip(health_row, "Archivo: pendiente")
        self.health_chip_file.grid(row=0, column=0, padx=(0, 6), sticky="ew")
        self.health_chip_exceptions = _build_chip(health_row, "Excepciones: pendientes")
        self.health_chip_exceptions.grid(row=0, column=1, padx=6, sticky="ew")
        self.health_chip_run = _build_chip(health_row, "Ejecucion: sin iniciar")
        self.health_chip_run.grid(row=0, column=2, padx=(6, 0), sticky="ew")

        kpi_grid = ctk.CTkFrame(main, fg_color="transparent")
        kpi_grid.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        for idx in range(5):
            kpi_grid.grid_columnconfigure(idx, weight=1)

        def _add_kpi(col: int, title: str, key: str) -> ctk.CTkLabel:
            card = ctk.CTkFrame(
                kpi_grid,
                corner_radius=12,
                border_width=1,
                border_color=COLOR_BORDER,
                fg_color=COLOR_PANEL_SOFT,
            )
            card.grid(row=0, column=col, padx=(0 if col == 0 else 6, 0), sticky="ew")
            accent = ctk.CTkFrame(card, corner_radius=8, height=6, fg_color=COLOR_KPI_NEUTRAL)
            accent.pack(fill="x", padx=10, pady=(8, 4))
            ctk.CTkLabel(card, text=title.upper(), text_color=COLOR_TEXT_MUTED, font=self.font_body).pack(
                anchor="w", padx=10, pady=(0, 0)
            )
            value = ctk.CTkLabel(
                card,
                text="-",
                text_color=COLOR_TEXT,
                font=ctk.CTkFont(family="Bahnschrift", size=22, weight="bold"),
            )
            value.pack(anchor="w", padx=10, pady=(2, 10))
            self._kpi_cards[key] = {"card": card, "accent": accent, "value": value}
            return value

        self.kpi_employees_value = _add_kpi(0, "Empleados", "employees")
        self.kpi_days_value = _add_kpi(1, "Jornadas", "days")
        self.kpi_late_value = _add_kpi(2, "Tardanzas", "late")
        self.kpi_absent_value = _add_kpi(3, "Ausencias", "absent")
        self.kpi_overtime_value = _add_kpi(4, "Horas Extra", "overtime")

        context_card = ctk.CTkFrame(
            main,
            corner_radius=12,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL_SOFT,
        )
        context_card.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        context_card.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(context_card, text="Archivo:", text_color=COLOR_TEXT_MUTED, font=self.font_body).grid(
            row=0, column=0, padx=(12, 8), pady=(10, 4), sticky="w"
        )
        self.context_file_value = ctk.CTkLabel(context_card, text="Sin archivo", text_color=COLOR_TEXT, font=self.font_body)
        self.context_file_value.grid(row=0, column=1, padx=(0, 12), pady=(10, 4), sticky="w")

        ctk.CTkLabel(context_card, text="Excepciones:", text_color=COLOR_TEXT_MUTED, font=self.font_body).grid(
            row=1, column=0, padx=(12, 8), pady=4, sticky="w"
        )
        self.context_exceptions_value = ctk.CTkLabel(context_card, text="Sin archivo", text_color=COLOR_TEXT, font=self.font_body)
        self.context_exceptions_value.grid(row=1, column=1, padx=(0, 12), pady=4, sticky="w")

        ctk.CTkLabel(context_card, text="Manual pendiente:", text_color=COLOR_TEXT_MUTED, font=self.font_body).grid(
            row=2, column=0, padx=(12, 8), pady=4, sticky="w"
        )
        self.context_manual_value = ctk.CTkLabel(context_card, text="0 lineas", text_color=COLOR_TEXT, font=self.font_body)
        self.context_manual_value.grid(row=2, column=1, padx=(0, 12), pady=4, sticky="w")

        ctk.CTkLabel(context_card, text="Ultimo reporte:", text_color=COLOR_TEXT_MUTED, font=self.font_body).grid(
            row=3, column=0, padx=(12, 8), pady=(4, 10), sticky="w"
        )
        self.context_output_value = ctk.CTkLabel(context_card, text="No generado", text_color=COLOR_TEXT, font=self.font_body)
        self.context_output_value.grid(row=3, column=1, padx=(0, 12), pady=(4, 10), sticky="w")

        self.log_box = ctk.CTkTextbox(main, corner_radius=10, border_width=1)
        self.log_box.configure(border_color=COLOR_BORDER, fg_color=COLOR_PANEL_SOFT, text_color=COLOR_TEXT, font=self.font_mono)
        self.log_box.grid(row=4, column=0, sticky="nsew")
        self._log_text_widget = getattr(self.log_box, "_textbox", None)
        if isinstance(self._log_text_widget, tk.Text):
            self._log_text_widget.tag_config("info", foreground=self._theme_color(COLOR_LOG_INFO))
            self._log_text_widget.tag_config("success", foreground=self._theme_color(COLOR_LOG_SUCCESS))
            self._log_text_widget.tag_config("warn", foreground=self._theme_color(COLOR_LOG_WARN))
            self._log_text_widget.tag_config("error", foreground=self._theme_color(COLOR_LOG_ERROR))
            self._log_text_widget.tag_config("time", foreground=self._theme_color(COLOR_TEXT_MUTED))
        self.log("Listo para procesar fichadas.")
        self.log_box.configure(state="disabled")

        watermark = ctk.CTkLabel(
            page,
            text="PORTA",
            font=ctk.CTkFont(size=116, weight="bold"),
            text_color=("#DCE6F2", "#172436"),
        )
        watermark.place(relx=0.64, rely=0.52, anchor="center")
        watermark.lower()

        return page

    def _build_exceptions_page(self, parent: ctk.CTkFrame) -> ctk.CTkFrame:
        page = ctk.CTkFrame(parent, corner_radius=0, fg_color="transparent")
        page.grid_columnconfigure(0, weight=0)
        page.grid_columnconfigure(1, weight=1)
        page.grid_rowconfigure(0, weight=1)

        sidebar = ctk.CTkFrame(
            page,
            width=370,
            corner_radius=18,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL,
        )
        sidebar.grid(row=0, column=0, sticky="nsw", padx=(14, 10), pady=12)
        sidebar.grid_propagate(False)
        sidebar.grid_columnconfigure(0, weight=1)

        title = ctk.CTkLabel(
            sidebar,
            text="Excepciones",
            font=self.font_title,
            text_color=COLOR_TEXT,
        )
        title.grid(row=0, column=0, padx=20, pady=(22, 8), sticky="w")

        subtitle = ctk.CTkLabel(
            sidebar,
            text="Configura feriados, vacaciones,\nenfermedad y permisos.",
            justify="left",
            text_color=COLOR_TEXT_MUTED,
            font=self.font_subtitle,
        )
        subtitle.grid(row=1, column=0, padx=20, pady=(0, 16), sticky="w")

        file_label = ctk.CTkLabel(
            sidebar,
            text="Archivo de excepciones",
            font=self.font_button,
            text_color=COLOR_TEXT,
        )
        file_label.grid(row=2, column=0, padx=20, pady=(0, 8), sticky="w")

        self.exceptions_file_entry = ctk.CTkEntry(sidebar, placeholder_text="Sin archivo de excepciones")
        self.exceptions_file_entry.grid(row=3, column=0, padx=20, pady=(0, 8), sticky="ew")
        self.exceptions_file_entry.configure(border_color=COLOR_BORDER, fg_color=COLOR_PANEL_SOFT, text_color=COLOR_TEXT)
        self.exceptions_file_entry.configure(state="disabled")

        self.select_exceptions_button = ctk.CTkButton(
            sidebar,
            text="Seleccionar archivo",
            height=34,
            command=self.select_exceptions_file,
        )
        self._style_primary_button(self.select_exceptions_button, height=34)
        self.select_exceptions_button.grid(row=4, column=0, padx=20, pady=(0, 8), sticky="ew")

        self.clear_exceptions_button = ctk.CTkButton(
            sidebar,
            text="Restaurar archivo por defecto",
            height=34,
            command=self.clear_exceptions_file,
        )
        self._style_secondary_button(self.clear_exceptions_button, height=34)
        self.clear_exceptions_button.grid(row=5, column=0, padx=20, pady=(0, 8), sticky="ew")

        self.save_exceptions_button = ctk.CTkButton(
            sidebar,
            text="Guardar configuracion",
            height=36,
            command=self.save_exceptions_config,
        )
        self._style_primary_button(self.save_exceptions_button, height=36)
        self.save_exceptions_button.grid(row=6, column=0, padx=20, pady=(6, 8), sticky="ew")

        self.open_date_button_exceptions = ctk.CTkButton(
            sidebar,
            text="Abrir date.xlsx",
            height=34,
            command=self.open_date_template_file,
        )
        self._style_secondary_button(self.open_date_button_exceptions, height=34)
        self.open_date_button_exceptions.grid(row=7, column=0, padx=20, pady=(0, 8), sticky="ew")

        info = ctk.CTkLabel(
            sidebar,
            text="Formatos: .csv/.xls/.xlsx",
            text_color=COLOR_TEXT_MUTED,
            justify="left",
            font=self.font_body,
        )
        info.grid(row=8, column=0, padx=20, pady=(0, 12), sticky="w")

        main = ctk.CTkFrame(
            page,
            corner_radius=18,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL,
        )
        main.grid(row=0, column=1, sticky="nsew", padx=(8, 14), pady=12)
        main.grid_columnconfigure(0, weight=1)
        main.grid_rowconfigure(4, weight=1)

        hero = ctk.CTkFrame(
            main,
            corner_radius=12,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL_SOFT,
        )
        hero.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        hero.grid_columnconfigure(0, weight=1)

        hero_accent = ctk.CTkFrame(hero, corner_radius=8, height=6, fg_color=COLOR_PRIMARY)
        hero_accent.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 0))

        header = ctk.CTkLabel(
            hero,
            text="Centro de Excepciones",
            font=self.font_section,
            text_color=COLOR_TEXT,
        )
        header.grid(row=1, column=0, sticky="w", padx=12, pady=(8, 2))

        helper = ctk.CTkLabel(
            hero,
            text=(
                "Carga rapida por campos o pega varias lineas. "
                "Rango inicio/fin ideal para vacaciones y licencias."
            ),
            justify="left",
            text_color=COLOR_TEXT_MUTED,
            font=self.font_subtitle,
        )
        helper.grid(row=2, column=0, sticky="w", padx=12, pady=(0, 8))

        hero_chips = ctk.CTkFrame(hero, fg_color="transparent")
        hero_chips.grid(row=3, column=0, sticky="ew", padx=12, pady=(0, 10))
        hero_chips.grid_columnconfigure((0, 1, 2), weight=1)

        def _build_exceptions_chip(parent: ctk.CTkFrame, text: str) -> ctk.CTkLabel:
            chip = ctk.CTkLabel(
                parent,
                text=text,
                corner_radius=8,
                fg_color=COLOR_CHIP_NEUTRAL_BG,
                text_color=COLOR_CHIP_NEUTRAL_TEXT,
                padx=8,
                pady=5,
                font=self.font_body,
            )
            return chip

        self.exceptions_chip_file = _build_exceptions_chip(hero_chips, "Archivo: sin archivo")
        self.exceptions_chip_file.grid(row=0, column=0, padx=(0, 6), sticky="ew")
        self.exceptions_chip_manual = _build_exceptions_chip(hero_chips, "Pendiente manual: 0 lineas")
        self.exceptions_chip_manual.grid(row=0, column=1, padx=6, sticky="ew")
        self.exceptions_chip_save = _build_exceptions_chip(hero_chips, "Guardado: disponible")
        self.exceptions_chip_save.grid(row=0, column=2, padx=(6, 0), sticky="ew")

        quick_card = ctk.CTkFrame(
            main,
            corner_radius=12,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL_SOFT,
        )
        quick_card.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        quick_card.grid_columnconfigure(0, weight=1)
        quick_card.grid_columnconfigure(1, weight=1)
        quick_card.grid_columnconfigure(2, weight=1)
        quick_card.grid_columnconfigure(3, weight=2)

        quick_title = ctk.CTkLabel(
            quick_card,
            text="Carga Rapida",
            font=self.font_button,
            text_color=COLOR_TEXT,
        )
        quick_title.grid(row=0, column=0, columnspan=4, sticky="w", padx=12, pady=(10, 8))

        self.quick_employee_menu = ctk.CTkOptionMenu(
            quick_card,
            values=self._employee_selector_values,
        )
        self.quick_employee_menu.configure(
            fg_color=COLOR_PANEL,
            button_color=COLOR_PRIMARY,
            button_hover_color=COLOR_PRIMARY_HOVER,
            text_color=COLOR_TEXT,
            font=self.font_body,
        )
        self.quick_employee_menu.grid(row=1, column=0, padx=(12, 6), pady=(0, 8), sticky="ew")
        self.quick_employee_menu.set("Todos")

        date_row = ctk.CTkFrame(quick_card, fg_color="transparent")
        date_row.grid(row=1, column=1, padx=6, pady=(0, 8), sticky="ew")
        date_row.grid_columnconfigure(0, weight=1)
        date_row.grid_columnconfigure(1, weight=1)
        date_row.grid_columnconfigure(2, weight=0)

        self.quick_date_entry = ctk.CTkEntry(
            date_row,
            placeholder_text="Inicio (YYYY-MM-DD o DD/MM/YYYY)",
        )
        self.quick_date_entry.configure(border_color=COLOR_BORDER, fg_color=COLOR_PANEL, text_color=COLOR_TEXT)
        self.quick_date_entry.grid(row=0, column=0, sticky="ew")

        self.quick_date_end_entry = ctk.CTkEntry(
            date_row,
            placeholder_text="Fin (opcional)",
        )
        self.quick_date_end_entry.configure(border_color=COLOR_BORDER, fg_color=COLOR_PANEL, text_color=COLOR_TEXT)
        self.quick_date_end_entry.grid(row=0, column=1, padx=(6, 0), sticky="ew")

        self.quick_date_button = ctk.CTkButton(
            date_row,
            text="Calendario",
            width=96,
            height=28,
            command=self.open_date_picker,
        )
        self._style_secondary_button(self.quick_date_button, height=28)
        self.quick_date_button.grid(row=0, column=2, padx=(6, 0), sticky="e")

        self.quick_type_menu = ctk.CTkOptionMenu(
            quick_card,
            values=self._absence_type_values,
        )
        self.quick_type_menu.configure(
            fg_color=COLOR_PANEL,
            button_color=COLOR_PRIMARY,
            button_hover_color=COLOR_PRIMARY_HOVER,
            text_color=COLOR_TEXT,
            font=self.font_body,
        )
        self.quick_type_menu.grid(row=1, column=2, padx=6, pady=(0, 8), sticky="ew")
        self.quick_type_menu.set(self._absence_type_values[0] if self._absence_type_values else "Feriado")

        self.quick_detail_entry = ctk.CTkEntry(quick_card, placeholder_text="Detalle (opcional)")
        self.quick_detail_entry.configure(border_color=COLOR_BORDER, fg_color=COLOR_PANEL, text_color=COLOR_TEXT)
        self.quick_detail_entry.grid(row=1, column=3, padx=(6, 12), pady=(0, 8), sticky="ew")

        self.quick_add_button = ctk.CTkButton(
            quick_card,
            text="Agregar a carga manual",
            height=32,
            command=self.add_manual_exception_from_fields,
        )
        self._style_primary_button(self.quick_add_button, height=32)
        self.quick_add_button.grid(row=2, column=0, columnspan=2, padx=(12, 6), pady=(0, 10), sticky="ew")

        self.quick_remove_last_button = ctk.CTkButton(
            quick_card,
            text="Quitar ultima linea",
            height=32,
            command=self.remove_last_manual_exception_line,
        )
        self._style_secondary_button(self.quick_remove_last_button, height=32)
        self.quick_remove_last_button.grid(row=2, column=2, padx=6, pady=(0, 10), sticky="ew")

        self.quick_clear_manual_button = ctk.CTkButton(
            quick_card,
            text="Limpiar manual",
            height=32,
            command=self.clear_manual_exceptions_box,
        )
        self._style_secondary_button(self.quick_clear_manual_button, height=32)
        self.quick_clear_manual_button.grid(row=2, column=3, padx=(6, 12), pady=(0, 10), sticky="ew")

        legend_card = ctk.CTkFrame(
            main,
            corner_radius=12,
            border_width=1,
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL_SOFT,
        )
        legend_card.grid(row=3, column=0, sticky="ew", pady=(0, 10))
        legend_card.grid_columnconfigure((0, 1, 2, 3), weight=1)

        ctk.CTkLabel(
            legend_card,
            text="Leyenda Rapida de Estados",
            font=self.font_button,
            text_color=COLOR_TEXT,
        ).grid(row=0, column=0, columnspan=4, padx=12, pady=(8, 6), sticky="w")

        legend_items = [
            ("Domingo", "#D9D9D9"),
            ("Ausencia", "#EF9A9A"),
            ("Vacaciones", "#FFF59D"),
            ("Estudiar", "#B3E5FC"),
            ("Capacitaciones", "#F8BBD0"),
            ("Suspencion", "#90CAF9"),
            ("No trabajado", "#D7CCC8"),
            ("Licencia", "#FFCC80"),
            ("Feriado", "#A9DF8F"),
            ("Accidente", "#CE93D8"),
            ("Tardanza", "#2E7D32"),
        ]
        for idx, (name, color) in enumerate(legend_items):
            row = 1 + (idx // 4)
            col = idx % 4
            chip = ctk.CTkFrame(legend_card, corner_radius=8, fg_color=color)
            chip.grid(row=row, column=col, padx=8, pady=6, sticky="ew")
            ctk.CTkLabel(
                chip,
                text=name,
                text_color="#102A43",
                font=self.font_body,
            ).pack(padx=8, pady=4)

        self.exceptions_manual_box = ctk.CTkTextbox(main, corner_radius=10, border_width=1)
        self.exceptions_manual_box.configure(
            border_color=COLOR_BORDER,
            fg_color=COLOR_PANEL_SOFT,
            text_color=COLOR_TEXT,
            font=self.font_mono,
        )
        self.exceptions_manual_box.grid(row=4, column=0, sticky="nsew")
        self.exceptions_manual_box.insert(
            "end",
            "# Formato por linea: ID|YYYY-MM-DD|TIPO|DETALLE\n"
            "# Ejemplo: 20|2026-05-01|Feriado|Dia del trabajador\n",
        )

        watermark = ctk.CTkLabel(
            page,
            text="PORTA",
            font=ctk.CTkFont(size=116, weight="bold"),
            text_color=("#DCE6F2", "#172436"),
        )
        watermark.place(relx=0.64, rely=0.52, anchor="center")
        watermark.lower()

        return page

    def _show_page(self, page: str, *, animate: bool = True) -> None:
        if page not in self._page_frames:
            return

        if page == "exceptions":
            self._refresh_absence_type_options()

        if page == self.current_page and not self._page_transition_running:
            self._apply_tab_visual_state(page)
            self._refresh_page_titles(page)
            return

        if not animate:
            if self._page_transition_after is not None:
                self.after_cancel(self._page_transition_after)
                self._page_transition_after = None
            self._page_transition_running = False
            self._pending_page = None
            target_frame = self._page_frames[page]
            for name, frame in self._page_frames.items():
                frame.place_configure(x=0, y=0, relwidth=1, relheight=1)
                if name == page:
                    frame.lift()
                else:
                    frame.lower()
            self.current_page = page
            self._apply_tab_visual_state(page)
            self._refresh_page_titles(page)
            return

        self._animate_page_switch(page)

    def _set_status(self, text: str, level: str = "ready") -> None:
        palette = {
            "ready": {"fg": COLOR_STATUS_READY_BG, "text": COLOR_STATUS_READY_TEXT},
            "working": {"fg": COLOR_STATUS_WORKING_BG, "text": COLOR_STATUS_WORKING_TEXT},
            "error": {"fg": COLOR_STATUS_ERROR_BG, "text": COLOR_STATUS_ERROR_TEXT},
        }
        colors = palette.get(level, palette["ready"])
        self.status_badge.configure(
            text=f"Estado: {text}",
            fg_color=colors["fg"],
            text_color=colors["text"],
        )
        if self.report_hero_subtitle is not None:
            self.report_hero_subtitle.configure(
                text=f"Estado actual: {text}. Bitacora en vivo del procesamiento y validaciones."
            )

        run_level_map = {"ready": "ok", "working": "working", "error": "error"}
        run_level = run_level_map.get(level, "neutral")
        self._set_chip_state(self.health_chip_run, f"Ejecucion: {text}", run_level)

        if level == "working":
            self._start_status_pulse()
        else:
            self._stop_status_pulse()

    def _set_file_entry(self, text: str) -> None:
        self.file_entry.configure(state="normal")
        self.file_entry.delete(0, "end")
        self.file_entry.insert(0, text)
        self.file_entry.configure(state="disabled")

    def _set_exceptions_file_entry(self, text: str) -> None:
        if self.exceptions_file_entry is None:
            return
        self.exceptions_file_entry.configure(state="normal")
        self.exceptions_file_entry.delete(0, "end")
        if text:
            self.exceptions_file_entry.insert(0, text)
        self.exceptions_file_entry.configure(state="disabled")

    def _set_processing_state(self, enabled: bool) -> None:
        self.select_button.configure(state="normal" if enabled else "disabled")
        self.process_button.configure(state="normal" if enabled and self.selected_file else "disabled")
        self.clear_button.configure(state="normal" if enabled else "disabled")
        if self.download_study_button is not None:
            self.download_study_button.configure(
                state="normal" if enabled and self.output_file is not None else "disabled"
            )

        self.select_exceptions_button.configure(state="normal" if enabled else "disabled")
        self.clear_exceptions_button.configure(state="normal" if enabled else "disabled")
        self.save_exceptions_button.configure(state="normal" if enabled else "disabled")
        if self.exceptions_manual_box is not None:
            self.exceptions_manual_box.configure(state="normal" if enabled else "disabled")
        if self.quick_employee_menu is not None:
            self.quick_employee_menu.configure(state="normal" if enabled else "disabled")
        if self.quick_date_entry is not None:
            self.quick_date_entry.configure(state="normal" if enabled else "disabled")
        if self.quick_date_end_entry is not None:
            self.quick_date_end_entry.configure(state="normal" if enabled else "disabled")
        if self.quick_date_button is not None:
            self.quick_date_button.configure(state="normal" if enabled else "disabled")
        if self.quick_type_menu is not None:
            self.quick_type_menu.configure(state="normal" if enabled else "disabled")
        if self.quick_detail_entry is not None:
            self.quick_detail_entry.configure(state="normal" if enabled else "disabled")
        if self.quick_add_button is not None:
            self.quick_add_button.configure(state="normal" if enabled else "disabled")
        if self.quick_remove_last_button is not None:
            self.quick_remove_last_button.configure(state="normal" if enabled else "disabled")
        if self.quick_clear_manual_button is not None:
            self.quick_clear_manual_button.configure(state="normal" if enabled else "disabled")
        if self.open_date_button_report is not None:
            self.open_date_button_report.configure(state="normal" if enabled else "disabled")
        if self.open_date_button_exceptions is not None:
            self.open_date_button_exceptions.configure(state="normal" if enabled else "disabled")

        self.report_tab_button.configure(state="normal" if enabled else "disabled")
        self.exceptions_tab_button.configure(state="normal" if enabled else "disabled")
        self._refresh_exceptions_hero_chips()

    def _set_progress(self, percent: int, step_text: str) -> None:
        bounded = max(0, min(100, int(percent)))
        self.progress.set(bounded / 100)
        self.progress_text.configure(text=f"Progreso: {bounded}%")
        self.progress_step.configure(text=f"Paso: {step_text}")

    def _update_elapsed(self) -> None:
        if self._started_at is None:
            self.elapsed_text.configure(text="Tiempo: 00:00")
            return
        elapsed = int(time.perf_counter() - self._started_at)
        minutes, seconds = divmod(elapsed, 60)
        self.elapsed_text.configure(text=f"Tiempo: {minutes:02d}:{seconds:02d}")

    def _count_manual_exceptions(self, text: str) -> int:
        lines = [line.strip() for line in text.splitlines()]
        return sum(1 for line in lines if line and not line.startswith("#"))

    def _refresh_exceptions_summary(self) -> None:
        file_text = self.exceptions_file.name if self.exceptions_file else "sin archivo"
        manual_text = self.exceptions_manual_box.get("1.0", "end").strip() if self.exceptions_manual_box else ""
        manual_count = self._count_manual_exceptions(manual_text)
        if self.exceptions_summary_label is not None:
            self.exceptions_summary_label.configure(
                text=f"Excepciones: {file_text}\nCarga manual: {manual_count} lineas"
            )
        self._refresh_context_panel()

    def _initialize_default_exceptions_file(self) -> None:
        # El archivo de feriados por defecto ya no se usa; las excepciones se
        # cargan desde date.xlsx (hoja Ausencias). Este método se mantiene para
        # no romper los llamados existentes pero no hace nada.
        pass

    def _manual_text_from_box(self) -> str:
        return self.exceptions_manual_box.get("1.0", "end").strip() if self.exceptions_manual_box else ""

    def _reset_manual_box_to_template(self) -> None:
        if self.exceptions_manual_box is None:
            return
        self.exceptions_manual_box.delete("1.0", "end")
        self.exceptions_manual_box.insert(
            "end",
            "# Formato por linea: ID|YYYY-MM-DD|TIPO|DETALLE\n"
            "# Ejemplo: 20|2026-05-01|Feriado|Dia del trabajador\n",
        )
        self._refresh_exceptions_summary()

    def _persist_manual_exceptions(self) -> int:
        manual_text = self._manual_text_from_box()
        if self._count_manual_exceptions(manual_text) == 0:
            return 0

        if self.exceptions_file is None:
            raise ExceptionConfigError("No hay archivo de excepciones seleccionado.")

        added_count = append_manual_exceptions_to_file(self.exceptions_file, manual_text)
        self._reset_manual_box_to_template()
        return added_count

    def _set_employee_selector_values(self, labels: list[str]) -> None:
        if not labels:
            labels = ["Todos"]

        self._employee_selector_values = labels
        if self.quick_employee_menu is not None:
            self.quick_employee_menu.configure(values=labels)
            if "Todos" in labels:
                self.quick_employee_menu.set("Todos")
            else:
                self.quick_employee_menu.set(labels[0])

    def _normalize_label(self, text: object) -> str:
        ascii_text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
        return " ".join(ascii_text.strip().lower().replace("_", " ").split())

    def _clean_employee_id(self, value: object) -> str:
        if pd.isna(value):
            return ""
        text = str(value).strip()
        if text.lower() in {"", "nan", "none", "nat", "-"}:
            return ""
        if text.endswith(".0") and text[:-2].isdigit():
            return text[:-2]
        return text

    def _extract_employee_pairs(
        self,
        df: pd.DataFrame,
        id_aliases: list[str],
        name_aliases: list[str],
    ) -> list[tuple[str, str]]:
        if df.empty:
            return []

        columns = [str(col) for col in df.columns]
        normalized = {self._normalize_label(col): col for col in columns}

        def _find_col(aliases: list[str]) -> str | None:
            for alias in aliases:
                key = self._normalize_label(alias)
                if key in normalized:
                    return normalized[key]
            for col in columns:
                col_key = self._normalize_label(col)
                if any(self._normalize_label(alias) in col_key for alias in aliases):
                    return col
            return None

        id_col = _find_col(id_aliases)
        name_col = _find_col(name_aliases)
        if id_col is None:
            return []

        by_id: dict[str, str] = {}
        for _, row in df.iterrows():
            emp_id = self._clean_employee_id(row.get(id_col, ""))
            if not emp_id:
                continue
            name_value = ""
            if name_col is not None:
                raw_name = str(row.get(name_col, "")).strip()
                if raw_name.lower() not in {"", "nan", "none", "nat", "-"}:
                    name_value = raw_name
            if emp_id not in by_id or (not by_id[emp_id] and name_value):
                by_id[emp_id] = name_value

        pairs = [(emp_id, by_id[emp_id] or f"ID {emp_id}") for emp_id in by_id]
        pairs.sort(key=lambda item: (item[1].lower(), item[0]))
        return pairs

    def _apply_employee_pairs(self, pairs: list[tuple[str, str]]) -> None:
        self._employee_label_to_id = {"Todos": ""}
        labels = ["Todos"]
        used_labels: set[str] = set()

        for emp_id, emp_name in pairs:
            label = emp_name
            if label in used_labels:
                label = f"{emp_name} ({emp_id})"
            used_labels.add(label)
            self._employee_label_to_id[label] = emp_id
            labels.append(label)

        self._set_employee_selector_values(labels)

    def _load_absence_types_from_template(self) -> list[str]:
        root = get_app_base_dir()
        path = root / "date.xlsx"
        if not path.exists():
            return list(DEFAULT_ABSENCE_TYPES)
        try:
            with pd.ExcelFile(path) as excel:
                sheet_map = {self._normalize_label(name): name for name in excel.sheet_names}
            config_sheet = sheet_map.get(self._normalize_label("Config"))
            if config_sheet is None:
                return list(DEFAULT_ABSENCE_TYPES)
            cfg = pd.read_excel(path, sheet_name=config_sheet)
            if cfg.empty:
                return list(DEFAULT_ABSENCE_TYPES)

            columns = [str(col) for col in cfg.columns]
            normalized = {self._normalize_label(col): col for col in columns}
            col_name = normalized.get(self._normalize_label("Tipos de ausencia"))
            if col_name is None:
                for col in columns:
                    if "tipo" in self._normalize_label(col) and "ausenc" in self._normalize_label(col):
                        col_name = col
                        break
            if col_name is None:
                return list(DEFAULT_ABSENCE_TYPES)

            values: list[str] = []
            seen: set[str] = set()
            for raw in cfg[col_name].tolist():
                text = str(raw).strip()
                if text.lower() in {"", "nan", "none", "nat", "-"}:
                    continue
                key = text.upper()
                if key in seen:
                    continue
                seen.add(key)
                values.append(key)
            return values or list(DEFAULT_ABSENCE_TYPES)
        except Exception:
            return list(DEFAULT_ABSENCE_TYPES)

    def _load_absence_states_from_template(self) -> list[str]:
        root = get_app_base_dir()
        path = root / "date.xlsx"
        if not path.exists():
            return list(DEFAULT_ABSENCE_STATES)
        try:
            with pd.ExcelFile(path) as excel:
                sheet_map = {self._normalize_label(name): name for name in excel.sheet_names}
            config_sheet = sheet_map.get(self._normalize_label("Config"))
            if config_sheet is None:
                return list(DEFAULT_ABSENCE_STATES)
            cfg = pd.read_excel(path, sheet_name=config_sheet)
            if cfg.empty:
                return list(DEFAULT_ABSENCE_STATES)

            columns = [str(col) for col in cfg.columns]
            normalized = {self._normalize_label(col): col for col in columns}
            col_name = normalized.get(self._normalize_label("Estados posibles"))
            if col_name is None:
                for col in columns:
                    if "estado" in self._normalize_label(col):
                        col_name = col
                        break
            if col_name is None:
                return list(DEFAULT_ABSENCE_STATES)

            values: list[str] = []
            seen: set[str] = set()
            for raw in cfg[col_name].tolist():
                text = str(raw).strip()
                if text.lower() in {"", "nan", "none", "nat", "-"}:
                    continue
                key = text.upper()
                if key in seen:
                    continue
                seen.add(key)
                values.append(key)
            return values or list(DEFAULT_ABSENCE_STATES)
        except Exception:
            return list(DEFAULT_ABSENCE_STATES)

    def _refresh_absence_type_options(self) -> None:
        values = self._load_absence_types_from_template()
        self._absence_type_values = values
        if self.quick_type_menu is not None:
            current = self.quick_type_menu.get().strip()
            self.quick_type_menu.configure(values=values)
            if current in values:
                self.quick_type_menu.set(current)
            elif values:
                self.quick_type_menu.set(values[0])

    def _sync_absence_types_dropdown_in_template(self) -> None:
        root = get_app_base_dir()
        path = root / "date.xlsx"
        if not path.exists():
            return
        values = self._absence_type_values or list(DEFAULT_ABSENCE_TYPES)
        state_values = self._load_absence_states_from_template()
        if not values:
            return
        try:
            wb = load_workbook(path)
            sheet_map = {self._normalize_label(name): name for name in wb.sheetnames}
            aus_name = sheet_map.get(self._normalize_label("Ausencias"))
            if aus_name is None:
                return
            ws = wb[aus_name]
            header_values = [self._normalize_label(cell.value) for cell in ws[1]]
            type_idx = None
            state_idx = None
            for idx, header in enumerate(header_values, start=1):
                if "tipo ausencia" in header or header == "tipo":
                    type_idx = idx
                if header == "estado" or "estado" in header:
                    state_idx = idx
            if type_idx is None and state_idx is None:
                return

            if type_idx is not None:
                type_col = get_column_letter(type_idx)
                formula_values = [v.replace('"', "'") for v in values]
                formula = '"' + ",".join(formula_values) + '"'
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f"{type_col}2:{type_col}5000")

            if state_idx is not None and state_values:
                state_col = get_column_letter(state_idx)
                state_formula_values = [v.replace('"', "'") for v in state_values]
                state_formula = '"' + ",".join(state_formula_values) + '"'
                state_dv = DataValidation(type="list", formula1=state_formula, allow_blank=True)
                ws.add_data_validation(state_dv)
                state_dv.add(f"{state_col}2:{state_col}5000")
            wb.save(path)
        except Exception:
            # Si el archivo esta abierto en Excel o no se puede guardar, lo ignoramos silenciosamente.
            return

    def _refresh_employee_options(self, source_path: Path | None = None) -> None:
        by_id: dict[str, str] = {}

        root = get_app_base_dir()
        base_candidates = [root / "date.xlsx", root / "info.xlsx"]
        for base_path in base_candidates:
            if not base_path.exists():
                continue
            try:
                sheet_to_use = None
                if base_path.suffix.lower() in {".xlsx", ".xls"}:
                    with pd.ExcelFile(base_path) as excel:
                        normalized_map = {
                            self._normalize_label(sheet_name): sheet_name for sheet_name in excel.sheet_names
                        }
                    sheet_to_use = normalized_map.get(self._normalize_label("Empleados"))
                if sheet_to_use is not None:
                    base_df = pd.read_excel(base_path, sheet_name=sheet_to_use)
                else:
                    base_df = pd.read_excel(base_path)

                base_pairs = self._extract_employee_pairs(
                    base_df,
                    id_aliases=["id", "legajo", "id de persona"],
                    name_aliases=["nombre", "name", "empleado"],
                )
                for emp_id, emp_name in base_pairs:
                    by_id[emp_id] = emp_name
                break
            except Exception as exc:
                self.log(f"No se pudo cargar lista de trabajadores desde {base_path.name}: {exc}")

        try:
            if source_path is not None:
                source_df = load_hikvision_excel(source_path)
                source_pairs = self._extract_employee_pairs(
                    source_df,
                    id_aliases=["employee_id", "id", "id de persona"],
                    name_aliases=["employee_name", "nombre", "name"],
                )
                for emp_id, emp_name in source_pairs:
                    if emp_id not in by_id:
                        by_id[emp_id] = emp_name
        except Exception as exc:
            self.log(f"No se pudo cargar lista de trabajadores para excepciones: {exc}")

        pairs = [(emp_id, by_id[emp_id]) for emp_id in by_id]
        pairs.sort(key=lambda item: (item[1].lower(), item[0]))
        self._apply_employee_pairs(pairs)

    def _load_employee_options_from_source(self, source_path: Path) -> None:
        self._refresh_employee_options(source_path)

    def open_date_picker(self) -> None:
        if self.quick_date_entry is None or self.quick_date_end_entry is None:
            return

        if Calendar is None:
            messagebox.showwarning(
                "Calendario no disponible",
                "Instala dependencia 'tkcalendar' para usar el selector visual de fecha.",
            )
            return

        seed_start = self._normalize_input_date(self.quick_date_entry.get().strip())
        seed_end = self._normalize_input_date(self.quick_date_end_entry.get().strip())
        seed_start_dt = datetime.strptime(seed_start, "%Y-%m-%d") if seed_start else datetime.now()
        seed_end_dt = datetime.strptime(seed_end, "%Y-%m-%d") if seed_end else seed_start_dt

        popup = ctk.CTkToplevel(self)
        popup.title("Seleccionar rango de fechas")
        popup.geometry("700x390")
        popup.resizable(False, False)
        popup.transient(self)
        popup.grab_set()

        body = ctk.CTkFrame(popup, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=12, pady=12)

        title_row = ctk.CTkFrame(body, fg_color="transparent")
        title_row.pack(fill="x", pady=(0, 6))
        ctk.CTkLabel(title_row, text="Inicio", font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=(46, 0))
        ctk.CTkLabel(title_row, text="Fin", font=ctk.CTkFont(size=12, weight="bold")).pack(side="right", padx=(0, 62))

        calendars_row = ctk.CTkFrame(body, fg_color="transparent")
        calendars_row.pack(fill="both", expand=True)
        calendars_row.grid_columnconfigure(0, weight=1)
        calendars_row.grid_columnconfigure(1, weight=1)

        calendar_start = Calendar(
            calendars_row,
            selectmode="day",
            year=seed_start_dt.year,
            month=seed_start_dt.month,
            day=seed_start_dt.day,
            date_pattern="yyyy-mm-dd",
        )
        calendar_start.grid(row=0, column=0, padx=(0, 8), sticky="nsew")

        calendar_end = Calendar(
            calendars_row,
            selectmode="day",
            year=seed_end_dt.year,
            month=seed_end_dt.month,
            day=seed_end_dt.day,
            date_pattern="yyyy-mm-dd",
        )
        calendar_end.grid(row=0, column=1, padx=(8, 0), sticky="nsew")

        actions = ctk.CTkFrame(body, fg_color="transparent")
        actions.pack(fill="x", pady=(10, 0))

        def _confirm() -> None:
            picked_start = calendar_start.get_date()
            picked_end = calendar_end.get_date()
            start_dt = datetime.strptime(picked_start, "%Y-%m-%d")
            end_dt = datetime.strptime(picked_end, "%Y-%m-%d")
            if end_dt < start_dt:
                messagebox.showwarning(
                    "Rango invalido",
                    "La fecha fin no puede ser menor a la fecha inicio.",
                    parent=popup,
                )
                return
            self.quick_date_entry.delete(0, "end")
            self.quick_date_entry.insert(0, picked_start)
            self.quick_date_end_entry.delete(0, "end")
            self.quick_date_end_entry.insert(0, picked_end)
            popup.destroy()

        ctk.CTkButton(
            actions,
            text="Cancelar",
            fg_color="transparent",
            border_width=1,
            text_color=("gray20", "gray90"),
            command=popup.destroy,
        ).pack(side="left")
        ctk.CTkButton(actions, text="Seleccionar", command=_confirm).pack(side="right")

    def _normalize_input_date(self, raw_value: str) -> str | None:
        text = raw_value.strip()
        if not text:
            return None

        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.strftime("%Y-%m-%d")
            except ValueError:
                continue

        return None

    def add_manual_exception_from_fields(self) -> None:
        if self.exceptions_manual_box is None:
            return

        selected_label = self.quick_employee_menu.get().strip() if self.quick_employee_menu else "Todos"
        employee = self._employee_label_to_id.get(selected_label, "")
        date_start_value = self.quick_date_entry.get().strip() if self.quick_date_entry else ""
        date_end_value = self.quick_date_end_entry.get().strip() if self.quick_date_end_entry else ""
        exception_type = self.quick_type_menu.get().strip() if self.quick_type_menu else ""
        details = self.quick_detail_entry.get().strip() if self.quick_detail_entry else ""

        normalized_start = self._normalize_input_date(date_start_value)
        if normalized_start is None:
            messagebox.showwarning(
                "Fecha invalida",
                "Fecha inicio invalida. Usa YYYY-MM-DD o DD/MM/YYYY.",
            )
            return

        normalized_end = normalized_start
        if date_end_value:
            normalized_end = self._normalize_input_date(date_end_value)
            if normalized_end is None:
                messagebox.showwarning(
                    "Fecha invalida",
                    "Fecha fin invalida. Usa YYYY-MM-DD o DD/MM/YYYY.",
                )
                return

        start_dt = datetime.strptime(normalized_start, "%Y-%m-%d")
        end_dt = datetime.strptime(normalized_end, "%Y-%m-%d")
        if end_dt < start_dt:
            messagebox.showwarning(
                "Rango invalido",
                "La fecha fin no puede ser menor a la fecha inicio.",
            )
            return

        if not exception_type:
            messagebox.showwarning("Tipo faltante", "Selecciona un tipo de excepcion.")
            return

        current = self.exceptions_manual_box.get("1.0", "end").rstrip()
        if current and not current.endswith("\n"):
            self.exceptions_manual_box.insert("end", "\n")
        generated_lines: list[str] = []
        cursor = start_dt
        while cursor <= end_dt:
            new_line = f"{employee}|{cursor.strftime('%Y-%m-%d')}|{exception_type}|{details}".rstrip()
            self.exceptions_manual_box.insert("end", new_line + "\n")
            generated_lines.append(new_line)
            cursor += timedelta(days=1)
        self.exceptions_manual_box.see("end")

        if self.quick_date_entry is not None:
            self.quick_date_entry.delete(0, "end")
        if self.quick_date_end_entry is not None:
            self.quick_date_end_entry.delete(0, "end")
        if self.quick_detail_entry is not None:
            self.quick_detail_entry.delete(0, "end")
        if self.quick_employee_menu is not None and "Todos" in self._employee_selector_values:
            self.quick_employee_menu.set("Todos")

        self._refresh_exceptions_summary()
        self.log(
            f"Excepcion agregada en rango: {normalized_start} a {normalized_end} "
            f"({len(generated_lines)} dias)."
        )

    def remove_last_manual_exception_line(self) -> None:
        if self.exceptions_manual_box is None:
            return

        lines = self.exceptions_manual_box.get("1.0", "end").splitlines()
        while lines and (not lines[-1].strip() or lines[-1].strip().startswith("#")):
            lines.pop()
        if not lines:
            return

        removed = lines.pop()
        rebuilt = "\n".join(lines)
        if rebuilt:
            rebuilt += "\n"
        self.exceptions_manual_box.delete("1.0", "end")
        self.exceptions_manual_box.insert("1.0", rebuilt)
        self._refresh_exceptions_summary()
        self.log(f"Se removio la ultima linea manual: {removed}")

    def clear_manual_exceptions_box(self) -> None:
        self._reset_manual_box_to_template()
        self.log("Carga manual de excepciones reiniciada.")

    def _push_progress(self, percent: int, message: str) -> None:
        self._progress_queue.put(("progress", (percent, message)))

    def _worker_run(self, source: Path, exceptions_file: Path | None, manual_text: str) -> None:
        try:
            output = self.processor.process_file(
                source,
                progress_callback=self._push_progress,
                exceptions_file=exceptions_file,
                manual_exceptions_text=manual_text,
            )
            self._progress_queue.put(("done", output))
        except Exception as exc:
            self._progress_queue.put(("error", exc))

    def _poll_worker(self) -> None:
        finished = False
        while not self._progress_queue.empty():
            event, payload = self._progress_queue.get_nowait()
            if event == "progress":
                percent, text = payload
                self._set_progress(int(percent), str(text))
                self.log(str(text))
            elif event == "done":
                output = Path(payload)
                self.output_file = output
                self._set_progress(100, "Proceso completado.")
                self._set_status("completado", "ready")
                self.log(f"Reporte generado: {output}")
                self._refresh_kpis_from_output(output)
                self.open_button.configure(state="normal")
                self._set_processing_state(True)
                self._refresh_context_panel()
                messagebox.showinfo("Proceso completado", f"Excel generado en:\n{output}")
                finished = True
            elif event == "error":
                self._set_status("error", "error")
                self.log(f"Error: {payload}")
                self._set_processing_state(True)
                self._refresh_context_panel()
                messagebox.showerror("Error", str(payload))
                finished = True

        self._update_elapsed()

        if finished:
            self._worker_thread = None
            self._started_at = None
            return

        if self._worker_thread and self._worker_thread.is_alive():
            self.after(120, self._poll_worker)
        else:
            self._worker_thread = None
            self._started_at = None
            self._set_processing_state(True)

    def log(self, message: str) -> None:
        if self.log_box is None:
            return
        timestamp = datetime.now().strftime("%H:%M:%S")
        level = self._log_level_for_message(message)
        line = f"[{timestamp}] {message}\n"
        self.log_box.configure(state="normal")
        if isinstance(self._log_text_widget, tk.Text):
            self._log_text_widget.insert("end", f"[{timestamp}] ", ("time",))
            self._log_text_widget.insert("end", f"{message}\n", (level,))
        else:
            self.log_box.insert("end", line)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def clear_log(self) -> None:
        if self.log_box is None:
            return
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        if isinstance(self._log_text_widget, tk.Text):
            now = datetime.now().strftime("%H:%M:%S")
            self._log_text_widget.insert("end", f"[{now}] ", ("time",))
            self._log_text_widget.insert("end", "Bitacora limpia.\n", ("info",))
        else:
            self.log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] Bitacora limpia.\n")
        self.log_box.configure(state="disabled")

    def open_output_file(self) -> None:
        if not self.output_file:
            return
        try:
            os.startfile(self.output_file)
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{exc}")

    def download_study_report(self) -> None:
        if not self.output_file:
            messagebox.showwarning("Atencion", "Primero genera un reporte.")
            return
        if not self.output_file.exists():
            messagebox.showerror("Error", f"No se encontro el archivo:\n{self.output_file}")
            return

        default_name = f"{self.output_file.stem}_estudio.xlsx"
        target_path = filedialog.asksaveasfilename(
            title="Guardar reporte estudio",
            defaultextension=".xlsx",
            initialfile=default_name,
            initialdir=str(self.output_file.parent),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not target_path:
            return

        try:
            workbook = load_workbook(self.output_file)
            target_sheet_name = None
            for sheet_name in workbook.sheetnames:
                if self._normalize_label(sheet_name) == self._normalize_label("hs-riorda"):
                    target_sheet_name = sheet_name
                    break
            if target_sheet_name is None:
                raise ValueError("No se encontro la hoja hs-riorda en el reporte generado.")

            for sheet_name in list(workbook.sheetnames):
                if sheet_name != target_sheet_name:
                    workbook.remove(workbook[sheet_name])
            self._apply_study_dynamic_formulas(workbook[target_sheet_name])
            workbook.save(target_path)
            self.log(f"Reporte estudio generado: {target_path}")
            messagebox.showinfo("Reporte estudio", f"Archivo generado en:\n{target_path}")
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo generar el reporte estudio:\n{exc}")

    def _apply_study_dynamic_formulas(self, worksheet) -> None:
        def _is_daily_row(row_idx: int) -> bool:
            value = worksheet.cell(row=row_idx, column=1).value
            return isinstance(value, (datetime, dt_date))

        def _label(row_idx: int) -> str:
            return str(worksheet.cell(row=row_idx, column=2).value or "").strip().lower()

        def _formula_for_cell_refs(refs: list[str]) -> str:
            if not refs:
                return "=0"
            return "=" + "+".join(refs)

        def _refs_for_rows(col_idx: int, rows: list[int]) -> list[str]:
            col = get_column_letter(col_idx)
            return [f"{col}{row}" for row in rows]

        employee_cols = list(range(4, worksheet.max_column + 1))
        week_rows: list[int] = []
        q1_week_common_rows: list[int] = []
        q2_week_common_rows: list[int] = []
        all_week_common_rows: list[int] = []
        q1_week_extra_rows: list[int] = []
        q2_week_extra_rows: list[int] = []
        all_week_extra_rows: list[int] = []
        current_week_half = "q1"
        last_week_rows: list[int] = []

        for row_idx in range(2, worksheet.max_row + 1):
            if _is_daily_row(row_idx):
                week_rows.append(row_idx)
                day_number = worksheet.cell(row=row_idx, column=3).value
                try:
                    day_int = int(day_number)
                except Exception:
                    day_int = 0
                if 1 <= day_int <= 15:
                    current_week_half = "q1"
                elif day_int > 15:
                    current_week_half = "q2"
                continue

            row_label = _label(row_idx)
            if row_label.startswith("total semana") and "horas comunes" in row_label:
                last_week_rows = list(week_rows)
                for col_idx in employee_cols:
                    refs = _refs_for_rows(col_idx, week_rows)
                    if not refs:
                        worksheet.cell(row=row_idx, column=col_idx).value = "=0"
                    else:
                        sum_formula = _formula_for_cell_refs(refs)[1:]
                        worksheet.cell(row=row_idx, column=col_idx).value = f"=MIN(44,{sum_formula})"
                if current_week_half == "q1":
                    q1_week_common_rows.append(row_idx)
                else:
                    q2_week_common_rows.append(row_idx)
                all_week_common_rows.append(row_idx)
                week_rows = []
            elif row_label.startswith("total semana") and "horas extras" in row_label:
                for col_idx in employee_cols:
                    refs = _refs_for_rows(col_idx, last_week_rows)
                    if not refs:
                        worksheet.cell(row=row_idx, column=col_idx).value = "=0"
                    else:
                        sum_formula = _formula_for_cell_refs(refs)[1:]
                        worksheet.cell(row=row_idx, column=col_idx).value = f"=MAX(0,{sum_formula}-44)"
                if current_week_half == "q1":
                    q1_week_extra_rows.append(row_idx)
                else:
                    q2_week_extra_rows.append(row_idx)
                all_week_extra_rows.append(row_idx)
                last_week_rows = []
                week_rows = []
            elif row_label == "1ra quincena - horas comunes":
                for col_idx in employee_cols:
                    col = get_column_letter(col_idx)
                    refs = [f"{col}{r}" for r in q1_week_common_rows]
                    worksheet.cell(row=row_idx, column=col_idx).value = _formula_for_cell_refs(refs)
                week_rows = []
            elif row_label == "1ra quincena - horas extras":
                for col_idx in employee_cols:
                    col = get_column_letter(col_idx)
                    refs = [f"{col}{r}" for r in q1_week_extra_rows]
                    worksheet.cell(row=row_idx, column=col_idx).value = _formula_for_cell_refs(refs)
                week_rows = []
            elif row_label == "2da quincena - horas comunes":
                for col_idx in employee_cols:
                    col = get_column_letter(col_idx)
                    refs = [f"{col}{r}" for r in q2_week_common_rows]
                    worksheet.cell(row=row_idx, column=col_idx).value = _formula_for_cell_refs(refs)
                week_rows = []
            elif row_label == "2da quincena - horas extras":
                for col_idx in employee_cols:
                    col = get_column_letter(col_idx)
                    refs = [f"{col}{r}" for r in q2_week_extra_rows]
                    worksheet.cell(row=row_idx, column=col_idx).value = _formula_for_cell_refs(refs)
                week_rows = []
            elif row_label == "total mes - horas comunes":
                for col_idx in employee_cols:
                    col = get_column_letter(col_idx)
                    refs = [f"{col}{r}" for r in all_week_common_rows]
                    worksheet.cell(row=row_idx, column=col_idx).value = _formula_for_cell_refs(refs)
            elif row_label == "total mes - horas extras":
                for col_idx in employee_cols:
                    col = get_column_letter(col_idx)
                    refs = [f"{col}{r}" for r in all_week_extra_rows]
                    worksheet.cell(row=row_idx, column=col_idx).value = _formula_for_cell_refs(refs)

    def open_date_template_file(self) -> None:
        date_path = get_app_base_dir() / "date.xlsx"
        if not date_path.exists():
            messagebox.showwarning(
                "Archivo no encontrado",
                f"No se encontro date.xlsx en:\n{date_path}",
            )
            return
        try:
            os.startfile(date_path)
            self.log(f"Abriendo plantilla: {date_path.name}")
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo abrir date.xlsx:\n{exc}")

    def select_file(self) -> None:
        filepath = filedialog.askopenfilename(
            title="Seleccionar exportacion Hikvision",
            filetypes=[("Excel", "*.xlsx *.xls")],
        )
        if not filepath:
            return

        self.selected_file = Path(filepath)
        self.output_file = None
        self.open_button.configure(state="disabled")
        if self.download_study_button is not None:
            self.download_study_button.configure(state="disabled")
        self._set_file_entry(str(self.selected_file))
        self._refresh_context_panel()

        stat = self.selected_file.stat()
        size_mb = stat.st_size / (1024 * 1024)
        modified = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")

        self.file_meta.configure(
            text=(
                f"Nombre: {self.selected_file.name}\n"
                f"Tamanio: {size_mb:.2f} MB\n"
                f"Modificado: {modified}"
            )
        )

        self.process_button.configure(state="normal")
        self._set_status("archivo cargado", "ready")
        self._set_progress(0, "Listo para iniciar.")
        self.elapsed_text.configure(text="Tiempo: 00:00")
        self._refresh_kpis_from_source(self.selected_file)
        self._load_employee_options_from_source(self.selected_file)
        self.log(f"Archivo seleccionado: {self.selected_file}")

    def select_exceptions_file(self) -> None:
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo de excepciones",
            filetypes=[("Excepciones", "*.csv *.xlsx *.xls")],
        )
        if not filepath:
            return

        selected = Path(filepath)
        try:
            self.exceptions_file = ensure_exceptions_file(selected)
        except ExceptionConfigError as exc:
            messagebox.showerror("Error", str(exc))
            return
        self._set_exceptions_file_entry(str(self.exceptions_file))
        self._refresh_exceptions_summary()

    def clear_exceptions_file(self) -> None:
        self._initialize_default_exceptions_file()
        self.log("Se restablecio el archivo de excepciones por defecto.")

    def save_exceptions_config(self) -> None:
        if self.exceptions_file is None:
            messagebox.showwarning("Atencion", "No hay archivo de excepciones seleccionado.")
            return
        try:
            added_count = self._persist_manual_exceptions()
        except ExceptionConfigError as exc:
            messagebox.showerror("Error", str(exc))
            return

        self._refresh_exceptions_summary()
        if added_count > 0:
            self.log(
                f"Configuracion guardada. Se agregaron {added_count} excepciones manuales en {self.exceptions_file.name}."
            )
        else:
            self.log("Configuracion guardada. No habia nuevas excepciones manuales para agregar.")

    def process_file(self) -> None:
        if not self.selected_file:
            messagebox.showwarning("Atencion", "Primero selecciona un archivo.")
            return
        if self._worker_thread and self._worker_thread.is_alive():
            return

        if self.exceptions_file is None:
            self._initialize_default_exceptions_file()

        try:
            added_count = self._persist_manual_exceptions()
        except ExceptionConfigError as exc:
            messagebox.showerror("Error", str(exc))
            return

        if added_count > 0 and self.exceptions_file is not None:
            self.log(
                f"Se guardaron {added_count} excepciones manuales en {self.exceptions_file.name} antes de procesar."
            )

        self.output_file = None
        self.open_button.configure(state="disabled")
        if self.download_study_button is not None:
            self.download_study_button.configure(state="disabled")
        self._refresh_context_panel()
        self._set_status("procesando", "working")
        self._set_progress(2, "Preparando ejecucion...")
        self._started_at = time.perf_counter()
        self._set_processing_state(False)
        self.log("Iniciando procesamiento...")

        self._worker_thread = threading.Thread(
            target=self._worker_run,
            args=(self.selected_file, self.exceptions_file, ""),
            daemon=True,
        )
        self._worker_thread.start()
        self.after(120, self._poll_worker)

    # ── EMPLOYEES PAGE ─────────────────────────────────────────────────────────
    # Column spec: (col_id, heading, width, anchor, df_col_name)
    _EMP_COLS = [
        ("id",          "Legajo",       70,  "center", "Id"),
        ("nombre",      "Nombre",       220, "w",      "Nombre"),
        ("ing_m",       "Ing. Mañana",  90,  "center", "horario ingreso Mañana"),
        ("sal_m",       "Sal. Mañana",  90,  "center", "Horario salida Mañana"),
        ("ing_t",       "Ing. Tarde",   90,  "center", "Horario Ingreso Tarde"),
        ("sal_t",       "Sal. Tarde",   90,  "center", "Horario Salida Tarde"),
        ("contrat",     "Contratación", 110, "center", "Contratacion"),
        ("corrido",     "H. Corrido",   90,  "center", "Horario corrido"),
        ("dias",        "Días",         60,  "center", "Dias"),
        ("grupo",       "Grupo",        70,  "center", "grupo"),
    ]

    def _build_employees_page(self, parent: ctk.CTkFrame) -> ctk.CTkFrame:
        import tkinter.ttk as ttk

        page = ctk.CTkFrame(parent, corner_radius=0, fg_color=COLOR_BG)
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(1, weight=1)

        # ── Header ──────────────────────────────────────────────────────────────
        header = ctk.CTkFrame(page, corner_radius=0, fg_color=COLOR_PANEL, height=56,
                               border_width=1, border_color=COLOR_BORDER)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_propagate(False)
        header.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(header, text="Gestión de Empleados",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color=COLOR_TEXT).grid(row=0, column=0, padx=18, pady=16, sticky="w")
        self._emp_status_label = ctk.CTkLabel(header, text="", font=self.font_body,
                                               text_color=COLOR_TEXT_MUTED)
        self._emp_status_label.grid(row=0, column=1, padx=12, sticky="e")

        # ── Body ─────────────────────────────────────────────────────────────────
        body = ctk.CTkFrame(page, corner_radius=0, fg_color="transparent")
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=12)
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(0, weight=1)

        # Treeview + scrollbars
        tree_frame = ctk.CTkFrame(body, corner_radius=10, fg_color=COLOR_PANEL,
                                   border_width=1, border_color=COLOR_BORDER)
        tree_frame.grid(row=0, column=0, sticky="nsew")
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Emp.Treeview",
                        background="#FFFFFF", foreground="#0F172A",
                        fieldbackground="#FFFFFF", rowheight=30,
                        font=("Segoe UI", 11))
        style.configure("Emp.Treeview.Heading",
                        background="#EEF3F8", foreground="#475569",
                        font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("Emp.Treeview",
                  background=[("selected", "#E0F2FE")],
                  foreground=[("selected", "#0C4A6E")])

        col_ids = [c[0] for c in self._EMP_COLS]
        self._emp_tree = ttk.Treeview(tree_frame, columns=col_ids, show="headings",
                                       style="Emp.Treeview", selectmode="browse")
        for col_id, heading, width, anchor, _ in self._EMP_COLS:
            self._emp_tree.heading(col_id, text=heading)
            self._emp_tree.column(col_id, width=width, minwidth=50, anchor=anchor)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=self._emp_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self._emp_tree.xview)
        self._emp_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._emp_tree.grid(row=0, column=0, sticky="nsew", padx=(6, 0), pady=(6, 0))
        vsb.grid(row=0, column=1, sticky="ns",  pady=(6, 0), padx=(0, 4))
        hsb.grid(row=1, column=0, sticky="ew",  padx=(6, 0), pady=(0, 4))

        # ── Buttons ──────────────────────────────────────────────────────────────
        btn_bar = ctk.CTkFrame(body, fg_color="transparent")
        btn_bar.grid(row=1, column=0, sticky="ew", pady=(10, 0))

        def _btn(text, cmd, primary=False):
            b = ctk.CTkButton(btn_bar, text=text, width=110, height=36,
                               corner_radius=8, font=self.font_button, command=cmd)
            if primary:
                self._style_primary_button(b)
            else:
                self._style_secondary_button(b)
            return b

        _btn("➕ Agregar",  self._emp_add).pack(side="left", padx=(0, 8))
        _btn("✏️ Editar",   self._emp_edit).pack(side="left", padx=(0, 8))
        _btn("🗑️ Eliminar", self._emp_delete).pack(side="left", padx=(0, 16))
        _btn("💾 Guardar",  self._emp_save, primary=True).pack(side="left")

        # ── Internal state ────────────────────────────────────────────────────────
        self._emp_df: "pd.DataFrame | None" = None
        self._emp_dirty = False

        page.bind("<Visibility>", lambda e: self._emp_load())
        self.after(200, self._emp_load)
        return page

    def _emp_date_path(self) -> "Path | None":
        base = get_app_base_dir()
        for candidate in [
            base / "date.xlsx",
            Path("date.xlsx"),
            base / "dist" / "date.xlsx",
        ]:
            if candidate.exists():
                return candidate
        return None

    def _emp_load(self) -> None:
        path = self._emp_date_path()
        if path is None:
            self._emp_status_label.configure(text="⚠ No se encontró date.xlsx")
            return
        try:
            df = pd.read_excel(path, sheet_name="Empleados", dtype=str)
            df = df.fillna("")
            self._emp_df = df
            self._emp_refresh_tree()
            self._emp_dirty = False
            self._emp_status_label.configure(
                text=f"Cargado: {path.name}  ({len(df)} empleados)")
        except Exception as exc:
            self._emp_status_label.configure(text=f"Error al leer: {exc}")

    def _emp_refresh_tree(self) -> None:
        if self._emp_df is None:
            return
        self._emp_tree.delete(*self._emp_tree.get_children())
        for _, row in self._emp_df.iterrows():
            vals = []
            for _, _, _, _, df_col in self._EMP_COLS:
                v = str(row.get(df_col, "")).strip()
                if v.lower() in {"nan", "none"}:
                    v = ""
                vals.append(v)
            self._emp_tree.insert("", "end", values=vals)

    # ── Row index helper ──────────────────────────────────────────────────────────
    def _emp_selected_index(self) -> "int | None":
        """Return 0-based df index of selected row, or None."""
        sel = self._emp_tree.selection()
        if not sel:
            return None
        return self._emp_tree.index(sel[0])

    def _emp_add(self) -> None:
        self._emp_open_dialog(mode="add")

    def _emp_edit(self) -> None:
        idx = self._emp_selected_index()
        if idx is None:
            messagebox.showwarning("Editar", "Seleccioná un empleado primero.")
            return
        row = self._emp_df.iloc[idx]
        current = {}
        for _, _, _, _, df_col in self._EMP_COLS:
            v = str(row.get(df_col, "")).strip()
            if v.lower() in {"nan", "none"}:
                v = ""
            current[df_col] = v
        self._emp_open_dialog(mode="edit", df_index=idx, current=current)

    def _emp_delete(self) -> None:
        idx = self._emp_selected_index()
        if idx is None:
            messagebox.showwarning("Eliminar", "Seleccioná un empleado primero.")
            return
        row = self._emp_df.iloc[idx]
        nombre = str(row.get("Nombre", "")).strip()
        emp_id = str(row.get("Id", "")).strip()
        if not messagebox.askyesno("Confirmar", f"¿Eliminar a {nombre} (ID {emp_id})?"):
            return
        self._emp_df = self._emp_df.drop(index=idx).reset_index(drop=True)
        self._emp_refresh_tree()
        self._emp_dirty = True
        self._emp_status_label.configure(text=f"Eliminado: {nombre} — guardá para confirmar")

    # ── Días helpers ──────────────────────────────────────────────────────────────
    _DIAS_LIST = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    _DIAS_KEYS  = ["lunes", "martes", "mi",        "jueves", "viernes", "sab",    "dom"]

    @staticmethod
    def _dias_parse(raw: str) -> set[str]:
        """Return set of day names (from _DIAS_LIST) present in raw string."""
        r = raw.lower()
        result: set[str] = set()
        # "lunes a viernes" → first 5 days
        if "a viernes" in r or "a vier" in r:
            result.update(["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"])
        if "a s" in r or "sabado" in r or "sábado" in r:
            result.update(["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"])
        # individual mentions
        if "lunes"     in r: result.add("Lunes")
        if "martes"    in r: result.add("Martes")
        if "mi"        in r and ("mier" in r or "miérc" in r or "miérc" in r or "miércoles" in r or "miercoles" in r): result.add("Miércoles")
        if "jueves"    in r: result.add("Jueves")
        if "viernes"   in r: result.add("Viernes")
        if "s" in r    and ("sab" in r or "sáb" in r): result.add("Sábado")
        if "domingo"   in r: result.add("Domingo")
        return result

    @staticmethod
    def _dias_format(selected: list[str]) -> str:
        """Convert selected day names to a natural-language string."""
        wdays = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
        ordered = [d for d in ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
                   if d in selected]
        if not ordered:
            return ""
        if ordered == wdays:
            return "lunes a viernes"
        if ordered == wdays + ["Sábado"]:
            return "lunes a sábado"
        if ordered == ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]:
            return "todos los días"
        # comma list
        low = [d.lower() for d in ordered]
        if len(low) == 1:
            return low[0]
        return ", ".join(low[:-1]) + " y " + low[-1]

    def _emp_open_dialog(self, mode: str, df_index: "int | None" = None,
                          current: "dict | None" = None) -> None:
        import tkinter as tk

        cur = current or {}
        dlg = ctk.CTkToplevel(self)
        dlg.title("Agregar empleado" if mode == "add" else "Editar empleado")
        dlg.geometry("440x620")
        dlg.resizable(False, True)
        dlg.grab_set()
        dlg.configure(fg_color=COLOR_PANEL)

        # Scrollable inner frame
        scroll = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=20, pady=(16, 4))

        entries: dict = {}

        # Collect existing grupo values for the combo
        existing_grupos: list[str] = []
        if self._emp_df is not None and "grupo" in self._emp_df.columns:
            vals = self._emp_df["grupo"].dropna().astype(str).str.strip()
            existing_grupos = sorted({v for v in vals if v and v.lower() not in {"nan","none"}})

        for _, heading, _, _, df_col in self._EMP_COLS:
            ctk.CTkLabel(scroll, text=heading, font=self.font_body,
                         text_color=COLOR_TEXT_MUTED, anchor="w").pack(fill="x", pady=(6, 1))
            default_val = cur.get(df_col, "")

            # ── Contratacion: fixed dropdown ────────────────────────────────────
            if df_col == "Contratacion":
                w = ctk.CTkComboBox(scroll, values=["", "Quincenal", "Mensual"],
                                    font=self.font_body, height=34, state="readonly")
                w.set(default_val)
                w.pack(fill="x")
                entries[df_col] = w

            # ── Horario corrido: SI / NO dropdown ──────────────────────────────
            elif df_col == "Horario corrido":
                w = ctk.CTkComboBox(scroll, values=["", "SI", "NO"],
                                    font=self.font_body, height=34, state="readonly")
                w.set(default_val.upper() if default_val.upper() in {"SI","NO"} else "")
                w.pack(fill="x")
                entries[df_col] = w

            # ── Dias: individual checkboxes ─────────────────────────────────────
            elif df_col == "Dias":
                checked = self._dias_parse(default_val)
                day_vars: dict[str, ctk.BooleanVar] = {}
                days_frame = ctk.CTkFrame(scroll, fg_color="#F1F5F9", corner_radius=8)
                days_frame.pack(fill="x", pady=(0, 2))
                for day in self._DIAS_LIST:
                    var = ctk.BooleanVar(value=(day in checked))
                    cb = ctk.CTkCheckBox(days_frame, text=day, variable=var,
                                         font=self.font_body, text_color=COLOR_TEXT,
                                         fg_color=COLOR_PRIMARY)
                    cb.pack(anchor="w", padx=12, pady=3)
                    day_vars[day] = var
                entries[df_col] = day_vars   # special: dict of BooleanVar

            # ── Grupo: editable combo with existing values ──────────────────────
            elif df_col == "grupo":
                w = ctk.CTkComboBox(scroll, values=existing_grupos or [""],
                                    font=self.font_body, height=34)
                w.set(default_val)
                w.pack(fill="x")
                entries[df_col] = w

            # ── Default: plain text entry ───────────────────────────────────────
            else:
                e = ctk.CTkEntry(scroll, font=self.font_body, height=34)
                e.pack(fill="x")
                if default_val:
                    e.insert(0, default_val)
                entries[df_col] = e

        def _get(df_col: str) -> str:
            widget = entries[df_col]
            if df_col == "Dias":
                # dict of BooleanVar
                selected = [d for d in self._DIAS_LIST if widget[d].get()]
                return self._dias_format(selected)
            return widget.get().strip()

        def _confirm():
            emp_id = _get("Id")
            nombre  = _get("Nombre")
            if not emp_id or not nombre:
                messagebox.showwarning("Datos incompletos",
                                       "Legajo y Nombre son obligatorios.", parent=dlg)
                return

            if self._emp_df is None:
                self._emp_df = pd.DataFrame(columns=[c[4] for c in self._EMP_COLS])

            new_vals = {df_col: _get(df_col) for _, _, _, _, df_col in self._EMP_COLS}

            if mode == "edit" and df_index is not None:
                for df_col, val in new_vals.items():
                    if df_col in self._emp_df.columns:
                        self._emp_df.at[df_index, df_col] = val
            else:
                row_dict = {col: "" for col in self._emp_df.columns}
                row_dict.update(new_vals)
                self._emp_df = pd.concat(
                    [self._emp_df, pd.DataFrame([row_dict])], ignore_index=True
                )

            self._emp_refresh_tree()
            self._emp_dirty = True
            self._emp_status_label.configure(text="Cambios pendientes - guarda para confirmar")
            dlg.destroy()

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(4, 14))
        cancel_btn = ctk.CTkButton(btn_row, text="Cancelar", width=100, height=34,
                                    font=self.font_button, command=dlg.destroy)
        self._style_secondary_button(cancel_btn)
        cancel_btn.pack(side="left")
        confirm_btn = ctk.CTkButton(btn_row, text="Confirmar", width=120, height=34,
                                     font=self.font_button, command=_confirm)
        self._style_primary_button(confirm_btn)
        confirm_btn.pack(side="right")

    def _emp_save(self) -> None:
        if self._emp_df is None:
            messagebox.showwarning("Guardar", "No hay datos cargados.")
            return
        path = self._emp_date_path()
        if path is None:
            messagebox.showerror("Guardar", "No se encontro date.xlsx.")
            return
        try:
            from openpyxl import load_workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.utils import get_column_letter

            # Column widths from the template
            _COL_WIDTHS = [10.0, 28.0, 18.0, 13.0, 13.0, 13.0, 18.0, 16.0, 22.3, 20.0]

            # Remove read-only attribute if set (Windows copies can set it)
            import os as _os
            try:
                _os.chmod(path, 0o644)
            except Exception:
                pass

            wb = load_workbook(path)
            ws = wb["Empleados"]

            # Read header from existing sheet to preserve column order
            n_cols = ws.max_column
            header = [ws.cell(1, c).value for c in range(1, n_cols + 1)]

            # Remove existing tables so we can re-add with correct ref
            for tbl_name in list(ws.tables.keys()):
                del ws.tables[tbl_name]

            # Delete all data rows (keep header row 1)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            # Write new data rows
            df = self._emp_df.copy().fillna("")
            for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
                for c_idx, col_name in enumerate(header, start=1):
                    val = str(row.get(col_name, "")).strip() if col_name else ""
                    if val.lower() in {"nan", "none"}:
                        val = ""
                    cell = ws.cell(r_idx, c_idx)
                    if col_name == "Id" and val.isdigit():
                        cell.value = int(val)
                    else:
                        cell.value = val or None

            # Rebuild Excel Table with correct ref
            n_data = len(df)
            last_row = 1 + n_data
            last_col = get_column_letter(n_cols)
            tbl = Table(displayName="TablaEmpleados", ref=f"A1:{last_col}{last_row}")
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tbl)

            # Restore column widths
            for i, w in enumerate(_COL_WIDTHS, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            try:
                wb.save(path)
            except PermissionError:
                # File locked (Excel open) - save to a temp copy instead
                from datetime import datetime
                ts = datetime.now().strftime("%H%M%S")
                tmp = path.parent / ("date_guardado_" + ts + ".xlsx")
                wb.save(tmp)
                messagebox.showwarning(
                    "Archivo en uso",
                    "date.xlsx esta abierto en Excel y no se pudo sobreescribir.\n\n"
                    "Se guardo una copia en:\n" + tmp.name + "\n\n"
                    "Cerra Excel y renombrala a date.xlsx."
                )
                self._emp_dirty = False
                self._emp_status_label.configure(
                    text="Guardado en copia temporal: " + tmp.name)
                return

            self._emp_dirty = False
            self._emp_status_label.configure(
                text=f"Guardado en {path.name}  ({n_data} empleados)")
        except Exception as exc:
            messagebox.showerror("Error al guardar", str(exc))
