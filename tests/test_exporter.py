from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from services.exporter import export_report


def _rgb(cell) -> str:
    return str(cell.fill.fgColor.rgb or "")


def test_diario_status_colors_fill_entire_row(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "30"],
            "Nombre": ["Ana", "Luis"],
            "Fecha": [pd.Timestamp("2026-05-01").date(), pd.Timestamp("2026-05-01").date()],
            "Departamento": ["A", "B"],
            "Estado": ["Tarde", "Ausente"],
            "Tramos trabajados": ["07:31 - 12:00", ""],
            "Minutos reales": [269, 0],
            "Minutos redondeados": [270, 0],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["04:30", "00:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Diario"]

    # Fila 2 = "Tarde" -> verde oscuro en toda la fila.
    assert _rgb(ws["A2"]).endswith("2E7D32")
    assert _rgb(ws["F2"]).endswith("2E7D32")

    # Fila 3 = "Ausente" -> rojo en toda la fila.
    assert _rgb(ws["A3"]).endswith("EF9A9A")
    assert _rgb(ws["J3"]).endswith("EF9A9A")


def test_diario_domingo_row_is_light_gray(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20"],
            "Nombre": ["Ana"],
            "Fecha": [pd.Timestamp("2026-05-10").date()],
            "Departamento": ["A"],
            "Estado": ["Domingo"],
            "Tramos trabajados": ["08:00 - 12:00"],
            "Minutos reales": [240],
            "Minutos redondeados": [240],
            "Minutos extra": [240],
            "Horas extra": ["04:00"],
            "Horas totales": ["04:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_domingo.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Diario"]
    assert _rgb(ws["A2"]).endswith("D9D9D9")
    assert _rgb(ws["G2"]).endswith("D9D9D9")


def test_diario_presente_row_has_no_status_color_override(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20"],
            "Nombre": ["Ana"],
            "Fecha": [pd.Timestamp("2026-05-20").date()],
            "Departamento": ["A"],
            "Estado": ["Presente"],
            "Tramos trabajados": [""],
            "Minutos reales": [540],
            "Minutos redondeados": [540],
            "Minutos extra": [0],
            "Horas extra": ["00:00"],
            "Horas totales": ["09:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_presente.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Diario"]
    # Fila par con estado sin color especifico: conserva alternado, no color de estado.
    assert _rgb(ws["A2"]).endswith("F5F8FC")


def test_export_does_not_create_resumen_estudio_sheet(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "30"],
            "Nombre": ["Ana", "Luis"],
            "Fecha": [pd.Timestamp("2026-05-05").date(), pd.Timestamp("2026-05-06").date()],
            "Departamento": ["A", "B"],
            "Estado": ["Normal", "Tarde"],
            "Tramos trabajados": ["07:30 - 12:00", "07:31 - 12:00"],
            "Minutos reales": [270, 269],
            "Minutos redondeados": [270, 270],
            "Minutos extra": [30, 0],
            "Horas extra": ["00:30", "00:00"],
            "Horas totales": ["04:30", "04:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_resumen.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    assert "resumen-estudio" not in wb.sheetnames


def test_export_creates_liquidar_sheet_with_status_colors(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "30", "20", "30"],
            "Nombre": ["Ana", "Luis", "Ana", "Luis"],
            "Fecha": [
                pd.Timestamp("2026-05-05").date(),
                pd.Timestamp("2026-05-05").date(),
                pd.Timestamp("2026-05-06").date(),
                pd.Timestamp("2026-05-06").date(),
            ],
            "Departamento": ["A", "B", "A", "B"],
            "Estado": ["Vacaciones", "Feriado", "Enfermedad", "Tardanza"],
            "Tramos trabajados": ["", "", "", "07:31 - 12:00"],
            "Minutos reales": [540, 540, 0, 270],
            "Minutos redondeados": [540, 540, 0, 270],
            "Minutos extra": [0, 0, 0, 0],
            "Horas extra": ["00:00", "00:00", "00:00", "00:00"],
            "Horas totales": ["09:00", "09:00", "00:00", "04:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_liquidar.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    assert "Liquidar" in wb.sheetnames
    ws = wb["Liquidar"]

    # Columnas base
    assert ws["A1"].value == "Fecha"
    assert ws["B1"].value == "Dia"
    assert ws["C1"].value == "Dia #"

    # Empleados esperados
    assert ws["D1"].value == "20 Ana"
    assert ws["E1"].value == "30 Luis"

    # 2026-05-05 row
    assert str(ws["A2"].value)[:10] == "2026-05-05"
    assert ws["D2"].value == 9
    assert ws["E2"].value == 9
    # Vacaciones amarillo / Feriado verde manzana
    assert _rgb(ws["D2"]).endswith("FFF59D")
    assert _rgb(ws["E2"]).endswith("A9DF8F")

    # 2026-05-06 row
    assert str(ws["A3"].value)[:10] == "2026-05-06"
    assert ws["D3"].value == 0
    assert ws["E3"].value == 4.5
    # Enfermedad naranja, tardanza verde oscuro.
    assert _rgb(ws["D3"]).endswith("FFCC80")
    assert _rgb(ws["E3"]).endswith("2E7D32")


def test_liquidar_uses_only_normal_hours_excluding_overtime(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20"],
            "Nombre": ["Ana"],
            "Fecha": [pd.Timestamp("2026-05-05").date()],
            "Departamento": ["A"],
            "Estado": ["Tardanza"],
            "Tramos trabajados": ["07:30 - 17:00"],
            "Minutos reales": [570],
            "Minutos redondeados": [570],   # 9.5h totales
            "Minutos extra": [90],          # 1.5h extra
            "Horas extra": ["01:30"],
            "Horas totales": ["09:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_liquidar_normales.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Liquidar"]

    # 9.5h - 1.5h = 8.0h normales
    assert ws["D2"].value == 8


def test_liquidar_uses_only_employees_from_date_xlsx_when_available(tmp_path: Path) -> None:
    empleados_df = pd.DataFrame(
        {
            "Id": ["20"],
            "Nombre": ["Ana"],
        }
    )
    with pd.ExcelWriter(tmp_path / "date.xlsx", engine="openpyxl") as writer:
        empleados_df.to_excel(writer, sheet_name="Empleados", index=False)

    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "30"],
            "Nombre": ["Ana", "Luis"],
            "Fecha": [pd.Timestamp("2026-05-05").date(), pd.Timestamp("2026-05-05").date()],
            "Departamento": ["A", "B"],
            "Estado": ["Normal", "Normal"],
            "Tramos trabajados": ["07:30 - 12:00", "08:00 - 12:00"],
            "Minutos reales": [270, 240],
            "Minutos redondeados": [270, 240],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["04:30", "04:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_liquidar_catalogo.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Liquidar"]

    # Solo debe incluir legajo 20 (de date.xlsx), no el 30.
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    assert "20 Ana" in headers
    assert "30 Luis" not in headers


def test_liquidar_includes_worked_saturday_rows(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "20"],
            "Nombre": ["Ana", "Ana"],
            "Fecha": [pd.Timestamp("2026-05-08").date(), pd.Timestamp("2026-05-09").date()],
            "Departamento": ["A", "A"],
            "Estado": ["Normal", "Tardanza"],
            "Tramos trabajados": ["07:30 - 12:00", "07:41-11:36 [07:30-11:30]"],
            "Minutos reales": [270, 235],
            "Minutos redondeados": [270, 240],
            "Minutos extra": [0, 240],
            "Horas extra": ["00:00", "04:00"],
            "Horas totales": ["04:30", "04:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_liquidar_sabado.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Liquidar"]

    dates = [str(ws.cell(row=row, column=1).value)[:10] for row in range(2, ws.max_row + 1)]
    assert "2026-05-09" in dates
    saturday_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value)[:10] == "2026-05-09":
            saturday_row = row
            break
    assert saturday_row is not None
    assert ws.cell(row=saturday_row, column=4).value == 4


def test_liquidar_saturday_exceptions_are_blank_per_employee(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "30"],
            "Nombre": ["Ana", "Luis"],
            "Fecha": [pd.Timestamp("2026-05-23").date(), pd.Timestamp("2026-05-23").date()],
            "Departamento": ["A", "B"],
            "Estado": ["Normal", "Accidente de trabajo"],
            "Tramos trabajados": ["07:30-11:30 [07:30-11:30]", ""],
            "Minutos reales": [240, 0],
            "Minutos redondeados": [240, 0],
            "Minutos extra": [240, 0],
            "Horas extra": ["04:00", "00:00"],
            "Horas totales": ["04:00", "00:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_liquidar_sabado_excepciones.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Liquidar"]

    assert str(ws["A2"].value)[:10] == "2026-05-23"
    # 20 Ana -> asistio sabado, mostrar horas trabajadas.
    assert ws["D2"].value == 4
    # 30 Luis -> sabado con excepcion, se elimina de la liquidacion del dia.
    assert ws["E2"].value in ("", None)


def test_liquidar_skips_saturday_when_no_attendance_punches(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "30"],
            "Nombre": ["Ana", "Luis"],
            "Fecha": [pd.Timestamp("2026-05-23").date(), pd.Timestamp("2026-05-23").date()],
            "Departamento": ["A", "B"],
            "Estado": ["Vacaciones", "Accidente de trabajo"],
            "Tramos trabajados": ["", ""],
            "Minutos reales": [0, 0],
            "Minutos redondeados": [0, 0],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["00:00", "00:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_liquidar_sabado_sin_fichadas.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Liquidar"]

    dates = [str(ws.cell(row=row, column=1).value)[:10] for row in range(2, ws.max_row + 1)]
    assert "2026-05-23" not in dates


def test_liquidar_adds_quincena_common_and_extra_totals(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "20", "20"],
            "Nombre": ["Ana", "Ana", "Ana"],
                "Fecha": [
                    pd.Timestamp("2026-05-05").date(),  # 1ra quincena
                    pd.Timestamp("2026-05-18").date(),  # 2da quincena
                    pd.Timestamp("2026-05-20").date(),  # 2da quincena
                ],
            "Departamento": ["A", "A", "A"],
            "Estado": ["Normal", "Tardanza", "Normal"],
            "Tramos trabajados": ["07:30 - 12:00", "07:30 - 17:30", "07:30 - 12:00"],
            "Minutos reales": [270, 600, 270],
            "Minutos redondeados": [270, 600, 270],
            "Minutos extra": [30, 120, 0],
            "Horas extra": ["00:30", "02:00", "00:00"],
            "Horas totales": ["04:30", "10:00", "04:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_liquidar_quincena.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Liquidar"]

    labels = {}
    for row in range(2, ws.max_row + 1):
        labels[str(ws.cell(row=row, column=2).value)] = row

    # 1ra quincena: comunes=4.0h, extras=0.5h
    r = labels["1ra Quincena - Horas Comunes"]
    assert ws.cell(row=r, column=4).value == 4.0
    r = labels["1ra Quincena - Horas Extras"]
    assert ws.cell(row=r, column=4).value == 0.5

    # 2da quincena: comunes=12.5h, extras=2.0h
    r = labels["2da Quincena - Horas Comunes"]
    assert ws.cell(row=r, column=4).value == 12.5
    r = labels["2da Quincena - Horas Extras"]
    assert ws.cell(row=r, column=4).value == 2.0

    # Totales mes: comunes=16.5h, extras=2.5h
    r = labels["Total Mes - Horas Comunes"]
    assert ws.cell(row=r, column=4).value == 16.5
    r = labels["Total Mes - Horas Extras"]
    assert ws.cell(row=r, column=4).value == 2.5

    # Orden esperado: cierra 1ra quincena, luego siguen dias de 2da quincena.
    row_q1_end = labels["1ra Quincena - Horas Extras"]
    row_day_18 = None
    row_day_20 = None
    for row in range(2, ws.max_row + 1):
        date_value = ws.cell(row=row, column=1).value
        text = str(date_value)
        if "2026-05-18" in text:
            row_day_18 = row
        if "2026-05-20" in text:
            row_day_20 = row
    assert row_day_18 is not None and row_day_18 > row_q1_end
    assert row_day_20 is not None and row_day_20 > row_q1_end


def test_liquidar_adds_status_legend_at_bottom(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "20"],
            "Nombre": ["Ana", "Ana"],
            "Fecha": [pd.Timestamp("2026-05-05").date(), pd.Timestamp("2026-05-06").date()],
            "Departamento": ["A", "A"],
            "Estado": ["Feriado", "Tardanza"],
            "Tramos trabajados": ["", "07:31-12:00 [07:30-12:00]"],
            "Minutos reales": [0, 269],
            "Minutos redondeados": [0, 240],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["00:00", "04:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_liquidar_legend.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["Liquidar"]

    title_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Leyenda de estados":
            title_row = row
            break
    assert title_row is not None

    found_feriado = False
    for row in range(title_row + 1, ws.max_row + 1):
        for text_col in (2, 4):
            if ws.cell(row=row, column=text_col).value == "Feriado":
                box_col = text_col - 1
                assert _rgb(ws.cell(row=row, column=box_col)).endswith("A9DF8F")
                found_feriado = True
    assert found_feriado


def test_export_creates_incidencias_sheet_with_summary_and_detail(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "20", "30", "30"],
            "Nombre": ["Ana", "Ana", "Luis", "Luis"],
            "Fecha": [
                pd.Timestamp("2026-05-05").date(),
                pd.Timestamp("2026-05-06").date(),
                pd.Timestamp("2026-05-05").date(),
                pd.Timestamp("2026-05-06").date(),
            ],
            "Departamento": ["A", "A", "B", "B"],
            "Estado": ["Tarde", "Ausente", "Normal", "Tardanza"],
            "Tramos trabajados": ["07:31 - 12:00", "", "08:00 - 12:00", "08:05 - 12:00"],
            "Minutos reales": [269, 0, 240, 235],
            "Minutos redondeados": [270, 0, 240, 240],
            "Minutos extra": [0, 0, 0, 0],
            "Horas extra": ["00:00", "00:00", "00:00", "00:00"],
            "Horas totales": ["04:30", "00:00", "04:00", "04:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_incidencias.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    assert "Incidencias" in wb.sheetnames
    ws = wb["Incidencias"]

    # Resumen (arriba)
    assert ws["A1"].value == "ID de persona"
    assert ws["B1"].value == "Nombre"
    assert ws["C1"].value == "Total tardanzas"
    assert ws["D1"].value == "Total ausencias"

    # Ana: 1 tarde, 1 ausente. Luis: 1 tardanza, 0 ausencias.
    summary = {}
    row = 2
    while ws.cell(row=row, column=1).value not in ("", None):
        emp_id = str(ws.cell(row=row, column=1).value)
        summary[emp_id] = {
            "tardanzas": ws.cell(row=row, column=3).value,
            "ausencias": ws.cell(row=row, column=4).value,
        }
        row += 1
    assert summary["20"]["tardanzas"] == 1
    assert summary["20"]["ausencias"] == 1
    assert summary["30"]["tardanzas"] == 1
    assert summary["30"]["ausencias"] == 0

    # Detalle debajo: fecha/estado/fichada de tarde+ausente.
    detail_header_row = row + 2
    assert ws.cell(row=detail_header_row, column=1).value == "ID de persona"
    assert ws.cell(row=detail_header_row, column=2).value == "Nombre"
    assert ws.cell(row=detail_header_row, column=3).value == "Fecha"
    assert ws.cell(row=detail_header_row, column=4).value == "Estado"
    assert ws.cell(row=detail_header_row, column=5).value == "Fichada"

    details = []
    current = detail_header_row + 1
    while ws.cell(row=current, column=1).value not in ("", None):
        fichada_value = ws.cell(row=current, column=5).value
        details.append(
            (
                str(ws.cell(row=current, column=1).value),
                str(ws.cell(row=current, column=4).value),
                "" if fichada_value in ("", None) else str(fichada_value),
            )
        )
        current += 1

    assert ("20", "Tarde", "07:31 - 12:00") in details
    assert ("20", "Ausente", "") in details
    assert ("30", "Tardanza", "08:05 - 12:00") in details


def test_export_creates_hs_riorda_sheet_with_only_target_employees(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "30"],
            "Nombre": ["DE CARLI GONZALO DAVID", "ANA TEST"],
            "Fecha": [pd.Timestamp("2026-05-05").date(), pd.Timestamp("2026-05-05").date()],
            "Departamento": ["A", "B"],
            "Estado": ["Normal", "Normal"],
            "Tramos trabajados": ["07:30 - 12:00", "08:00 - 12:00"],
            "Minutos reales": [270, 240],
            "Minutos redondeados": [270, 240],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["04:30", "04:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    assert "hs-riorda" in wb.sheetnames
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    assert "20 De Carli Gonzalo" in headers
    assert "67 Madera Adrián" in headers
    assert "30 ANA TEST" not in headers
    # 3 columnas base + 20 empleados definidos.
    assert ws.max_column == 23


def test_hs_riorda_excludes_saturday_and_extra_summary_rows(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20", "20"],
            "Nombre": ["De Carli Gonzalo", "De Carli Gonzalo"],
            "Fecha": [pd.Timestamp("2026-05-08").date(), pd.Timestamp("2026-05-09").date()],  # viernes y sabado
            "Departamento": ["A", "A"],
            "Estado": ["Normal", "Normal"],
            "Tramos trabajados": ["07:30 - 12:00", "07:30 - 11:30"],
            "Minutos reales": [270, 240],
            "Minutos redondeados": [270, 240],
            "Minutos extra": [0, 240],
            "Horas extra": ["00:00", "04:00"],
            "Horas totales": ["04:30", "04:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_sin_sabados.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    rows = [
        [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1)
    ]
    labels = [str(row[1] or "") for row in rows]

    assert "Sabado" not in labels
    assert "1ra Quincena - Horas Extras" not in labels
    assert "2da Quincena - Horas Extras" not in labels
    assert "Total Mes - Horas Extras" not in labels


def test_hs_riorda_caps_normal_hours_by_employee(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["83", "119"],
            "Nombre": ["Zurvera Mirna", "Lencina Rocio Pilar"],
            "Fecha": [pd.Timestamp("2026-05-11").date(), pd.Timestamp("2026-05-11").date()],
            "Departamento": ["A", "A"],
            "Estado": ["Normal", "Normal"],
            "Tramos trabajados": ["07:00 - 17:00", "07:00 - 16:00"],
            "Minutos reales": [600, 540],
            "Minutos redondeados": [600, 540],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["10:00", "09:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_topes.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): idx + 1 for idx, h in enumerate(headers)}

    mirna_col = header_to_idx["83 Zurvera Mirna"]
    rocio_col = header_to_idx["119 Lencina Rocio Pilar"]

    monday_row = None
    q1_total_row = None
    month_total_row = None
    for row_idx in range(2, ws.max_row + 1):
        dia = str(ws.cell(row=row_idx, column=2).value or "")
        if dia == "Lunes":
            monday_row = row_idx
        if dia == "1ra Quincena - Horas Comunes":
            q1_total_row = row_idx
        if dia == "Total Mes - Horas Comunes":
            month_total_row = row_idx

    assert monday_row is not None
    assert q1_total_row is not None
    assert month_total_row is not None

    assert ws.cell(row=monday_row, column=mirna_col).value == 7
    assert ws.cell(row=monday_row, column=rocio_col).value == 8
    assert ws.cell(row=q1_total_row, column=mirna_col).value == 7
    assert ws.cell(row=q1_total_row, column=rocio_col).value == 8
    assert ws.cell(row=month_total_row, column=mirna_col).value == 7
    assert ws.cell(row=month_total_row, column=rocio_col).value == 8


def test_hs_riorda_ignores_tardiness_discount(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20"],
            "Nombre": ["De Carli Gonzalo"],
            "Fecha": [pd.Timestamp("2026-05-11").date()],
            "Departamento": ["A"],
            "Estado": ["Tarde"],
            "Tramos trabajados": ["07:31 - 18:00"],
            "Minutos reales": [509],
            "Minutos redondeados": [510],  # 8.5h para simular descuento por tardanza
            "Minutos extra": [0],
            "Horas extra": ["00:00"],
            "Horas totales": ["08:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_tardanza.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): idx + 1 for idx, h in enumerate(headers)}
    employee_col = header_to_idx["20 De Carli Gonzalo"]

    monday_row = None
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_idx, column=2).value or "") == "Lunes":
            monday_row = row_idx
            break
    assert monday_row is not None
    # No descuenta por tardanza en hs-riorda: queda jornada normal de 9h.
    assert ws.cell(row=monday_row, column=employee_col).value == 9


def test_hs_riorda_never_uses_dark_green_fill(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20"],
            "Nombre": ["De Carli Gonzalo"],
            "Fecha": [pd.Timestamp("2026-05-12").date()],
            "Departamento": ["A"],
            "Estado": ["Tardanza"],
            "Tramos trabajados": ["07:31 - 18:00"],
            "Minutos reales": [509],
            "Minutos redondeados": [510],
            "Minutos extra": [0],
            "Horas extra": ["00:00"],
            "Horas totales": ["08:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_sin_verde_oscuro.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            rgb = str(getattr(getattr(cell.fill, "fgColor", None), "rgb", "") or "").upper()
            assert not rgb.endswith("2E7D32")


def test_hs_riorda_adds_weekly_totals_and_keeps_quincena_total(tmp_path: Path) -> None:
    dates = pd.date_range("2026-05-04", "2026-05-15", freq="B")  # dos semanas completas lun-vie
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20"] * len(dates),
            "Nombre": ["De Carli Gonzalo"] * len(dates),
            "Fecha": [d.date() for d in dates],
            "Departamento": ["A"] * len(dates),
            "Estado": ["Normal"] * len(dates),
            "Tramos trabajados": ["07:30 - 18:00"] * len(dates),
            "Minutos reales": [540] * len(dates),
            "Minutos redondeados": [540] * len(dates),
            "Minutos extra": [0] * len(dates),
            "Horas extra": ["00:00"] * len(dates),
            "Horas totales": ["09:00"] * len(dates),
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_semanas.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    labels = [str(ws.cell(row=r, column=2).value or "") for r in range(2, ws.max_row + 1)]

    assert "Total Semana 1 - Horas Comunes" in labels
    assert "Total Semana 2 - Horas Comunes" in labels
    assert "1ra Quincena - Horas Comunes" in labels

    first_friday_idx = labels.index("Viernes")
    week_1_idx = labels.index("Total Semana 1 - Horas Comunes")
    second_friday_idx = labels.index("Viernes", first_friday_idx + 1)
    week_2_idx = labels.index("Total Semana 2 - Horas Comunes")
    q1_idx = labels.index("1ra Quincena - Horas Comunes")

    # Debajo de cada viernes aparece el total semanal.
    assert week_1_idx == first_friday_idx + 1
    assert week_2_idx == second_friday_idx + 1
    # Luego se mantiene el total de quincena.
    assert q1_idx > week_2_idx


def test_hs_riorda_marks_missing_scheduled_weekday_as_absent_and_keeps_non_working_day_blank(tmp_path: Path) -> None:
    empleados_df = pd.DataFrame(
        {
            "Id": ["113", "118"],
            "Nombre": ["Pelliza Roque", "Zeller Lucas Ezequiel"],
            "horario ingreso Mañana": ["07:30:00", "07:30:00"],
            "Horario salida Mañana": ["12:00:00", "12:00:00"],
            "Horario Ingreso Tarde": ["13:30:00", None],
            "Horario Salida Tarde": ["18:00:00", None],
            "Horario corrido": ["NO", "SI"],
            "Dias": ["lunes a viernes", "lunes , miercoles y viernes"],
        }
    )
    with pd.ExcelWriter(tmp_path / "date.xlsx", engine="openpyxl") as writer:
        empleados_df.to_excel(writer, sheet_name="Empleados", index=False)

    daily_df = pd.DataFrame(
        {
            "ID de persona": ["113", "118"],
            "Nombre": ["Pelliza Roque", "Zeller Lucas Ezequiel"],
            "Fecha": [pd.Timestamp("2026-05-20").date(), pd.Timestamp("2026-05-22").date()],
            "Departamento": ["A", "A"],
            "Estado": ["Normal", "Normal"],
            "Tramos trabajados": ["07:30 - 18:00", "07:30 - 12:00"],
            "Minutos reales": [540, 270],
            "Minutos redondeados": [540, 270],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["09:00", "04:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_ausente_programado.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): idx + 1 for idx, h in enumerate(headers)}
    pelliza_col = header_to_idx["113 Pelliza Roque"]
    zeller_col = header_to_idx["118 Zeller Lucas Ezequiel"]

    row_2026_05_22 = None
    row_2026_05_21 = None
    for row_idx in range(2, ws.max_row + 1):
        date_text = str(ws.cell(row=row_idx, column=1).value)
        if "2026-05-22" in date_text:
            row_2026_05_22 = row_idx
        if "2026-05-21" in date_text:
            row_2026_05_21 = row_idx

    assert row_2026_05_22 is not None
    assert row_2026_05_21 is not None

    # Pelliza trabaja lun-vie: sin fichada en viernes debe verse como ausente 0.00.
    assert ws.cell(row=row_2026_05_22, column=pelliza_col).value == 0
    assert _rgb(ws.cell(row=row_2026_05_22, column=pelliza_col)).endswith("EF9A9A")
    # Zeller no trabaja jueves: si no hay fichada, debe quedar en blanco (no ausente).
    assert ws.cell(row=row_2026_05_21, column=zeller_col).value in ("", None)


def test_hs_riorda_marks_absent_for_all_except_zeller_even_if_days_config_differs(tmp_path: Path) -> None:
    empleados_df = pd.DataFrame(
        {
            "Id": ["113", "118"],
            "Nombre": ["Pelliza Roque", "Zeller Lucas Ezequiel"],
            "horario ingreso Mañana": ["07:30:00", "07:30:00"],
            "Horario salida Mañana": ["12:00:00", "12:00:00"],
            "Horario Ingreso Tarde": ["13:30:00", None],
            "Horario Salida Tarde": ["18:00:00", None],
            "Horario corrido": ["NO", "SI"],
            "Dias": ["lunes , miercoles y viernes", "lunes , miercoles y viernes"],
        }
    )
    with pd.ExcelWriter(tmp_path / "date.xlsx", engine="openpyxl") as writer:
        empleados_df.to_excel(writer, sheet_name="Empleados", index=False)

    # Solo cargamos registros de miercoles/viernes; jueves queda sin registro.
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["113", "113", "118"],
            "Nombre": ["Pelliza Roque", "Pelliza Roque", "Zeller Lucas Ezequiel"],
            "Fecha": [
                pd.Timestamp("2026-05-20").date(),
                pd.Timestamp("2026-05-22").date(),
                pd.Timestamp("2026-05-22").date(),
            ],
            "Departamento": ["A", "A", "A"],
            "Estado": ["Normal", "Normal", "Normal"],
            "Tramos trabajados": ["07:30 - 18:00", "07:30 - 18:00", "07:30 - 12:00"],
            "Minutos reales": [540, 540, 270],
            "Minutos redondeados": [540, 540, 270],
            "Minutos extra": [0, 0, 0],
            "Horas extra": ["00:00", "00:00", "00:00"],
            "Horas totales": ["09:00", "09:00", "04:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_ausente_general.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): idx + 1 for idx, h in enumerate(headers)}
    pelliza_col = header_to_idx["113 Pelliza Roque"]
    zeller_col = header_to_idx["118 Zeller Lucas Ezequiel"]

    row_2026_05_21 = None
    for row_idx in range(2, ws.max_row + 1):
        if "2026-05-21" in str(ws.cell(row=row_idx, column=1).value):
            row_2026_05_21 = row_idx
            break
    assert row_2026_05_21 is not None

    # Pelliza (no Zeller): jueves sin registro debe ir Ausente 0.00 en rojo.
    assert ws.cell(row=row_2026_05_21, column=pelliza_col).value == 0
    assert _rgb(ws.cell(row=row_2026_05_21, column=pelliza_col)).endswith("EF9A9A")
    # Zeller: jueves sin registro sigue en blanco.
    assert ws.cell(row=row_2026_05_21, column=zeller_col).value in ("", None)


def test_hs_riorda_hides_zeller_non_working_days_even_with_absent_record(tmp_path: Path) -> None:
    empleados_df = pd.DataFrame(
        {
            "Id": ["118"],
            "Nombre": ["Zeller Lucas Ezequiel"],
            "horario ingreso Mañana": ["07:30:00"],
            "Horario salida Mañana": ["12:00:00"],
            "Horario Ingreso Tarde": [None],
            "Horario Salida Tarde": [None],
            "Horario corrido": ["SI"],
            "Dias": ["lunes , miercoles y viernes"],
        }
    )
    with pd.ExcelWriter(tmp_path / "date.xlsx", engine="openpyxl") as writer:
        empleados_df.to_excel(writer, sheet_name="Empleados", index=False)

    # Simula que Diario trae un Ausente en jueves para Zeller.
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["118", "118"],
            "Nombre": ["Zeller Lucas Ezequiel", "Zeller Lucas Ezequiel"],
            "Fecha": [pd.Timestamp("2026-05-21").date(), pd.Timestamp("2026-05-22").date()],
            "Departamento": ["A", "A"],
            "Estado": ["Ausente", "Normal"],
            "Tramos trabajados": ["", "07:30 - 12:00"],
            "Minutos reales": [0, 270],
            "Minutos redondeados": [0, 270],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["00:00", "04:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_zeller_no_jueves.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): idx + 1 for idx, h in enumerate(headers)}
    zeller_col = header_to_idx["118 Zeller Lucas Ezequiel"]

    thursday_row = None
    for row_idx in range(2, ws.max_row + 1):
        if "2026-05-21" in str(ws.cell(row=row_idx, column=1).value):
            thursday_row = row_idx
            break

    assert thursday_row is not None
    assert ws.cell(row=thursday_row, column=zeller_col).value in ("", None)


def test_hs_riorda_total_rows_have_stronger_borders(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["20"] * 5,
            "Nombre": ["De Carli Gonzalo"] * 5,
            "Fecha": [d.date() for d in pd.date_range("2026-05-11", "2026-05-15", freq="B")],
            "Departamento": ["A"] * 5,
            "Estado": ["Normal"] * 5,
            "Tramos trabajados": ["07:30 - 18:00"] * 5,
            "Minutos reales": [540] * 5,
            "Minutos redondeados": [540] * 5,
            "Minutos extra": [0] * 5,
            "Horas extra": ["00:00"] * 5,
            "Horas totales": ["09:00"] * 5,
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_bordes_totales.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]

    total_row = None
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_idx, column=2).value or "") == "Total Semana 1 - Horas Comunes":
            total_row = row_idx
            break
    assert total_row is not None

    cell = ws.cell(row=total_row, column=1)
    assert cell.border.top.style == "medium"
    assert cell.border.bottom.style == "medium"


def test_hs_riorda_highlights_below_daily_target_in_orange(tmp_path: Path) -> None:
    dates = [pd.Timestamp("2026-05-11").date(), pd.Timestamp("2026-05-12").date()]
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["118", "118"],
            "Nombre": ["Zeller Lucas Ezequiel", "Zeller Lucas Ezequiel"],
            "Fecha": dates,
            "Departamento": ["A", "A"],
            "Estado": ["Normal", "Normal"],
            "Tramos trabajados": ["08:00 - 12:00", "07:30 - 12:00"],  # 4h y 4.5h (objetivo 4.5)
            "Minutos reales": [240, 270],
            "Minutos redondeados": [240, 270],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["04:00", "04:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_objetivo_diario.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): idx + 1 for idx, h in enumerate(headers)}
    employee_col = header_to_idx["118 Zeller Lucas Ezequiel"]

    monday_row = None
    tuesday_row = None
    for row_idx in range(2, ws.max_row + 1):
        label = str(ws.cell(row=row_idx, column=2).value or "")
        if label == "Lunes":
            monday_row = row_idx
        if label == "Martes":
            tuesday_row = row_idx
    assert monday_row is not None
    assert tuesday_row is not None

    monday_cell = ws.cell(row=monday_row, column=employee_col)
    monday_rgb = str(getattr(getattr(monday_cell.fill, "fgColor", None), "rgb", "") or "").upper()
    assert monday_rgb.endswith("FFB74D")

    tuesday_cell = ws.cell(row=tuesday_row, column=employee_col)
    tuesday_rgb = str(getattr(getattr(tuesday_cell.fill, "fgColor", None), "rgb", "") or "").upper()
    assert not tuesday_rgb.endswith("FFB74D")


def test_hs_riorda_does_not_discount_system_extra_for_capped_profiles(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["118", "119"],
            "Nombre": ["Zeller Lucas Ezequiel", "Lencina Rocio Pilar"],
            "Fecha": [pd.Timestamp("2026-05-13").date(), pd.Timestamp("2026-05-13").date()],
            "Departamento": ["A", "A"],
            "Estado": ["Normal", "Normal"],
            "Tramos trabajados": ["07:00 - 12:00", "07:00 - 15:00"],
            "Minutos reales": [300, 480],
            "Minutos redondeados": [300, 480],
            # El sistema base marca parte como extra, pero en hs-riorda deben contar normales hasta el tope.
            "Minutos extra": [60, 120],
            "Horas extra": ["01:00", "02:00"],
            "Horas totales": ["05:00", "08:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_topes_sin_descuento_extra.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): idx + 1 for idx, h in enumerate(headers)}
    zeller_col = header_to_idx["118 Zeller Lucas Ezequiel"]
    rocio_col = header_to_idx["119 Lencina Rocio Pilar"]

    wednesday_row = None
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_idx, column=2).value or "") == "Miercoles":
            wednesday_row = row_idx
            break
    assert wednesday_row is not None

    assert ws.cell(row=wednesday_row, column=zeller_col).value == 4.5
    assert ws.cell(row=wednesday_row, column=rocio_col).value == 8


def test_hs_riorda_preserves_status_colors_when_not_below_target(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["118"],
            "Nombre": ["Zeller Lucas Ezequiel"],
            "Fecha": [pd.Timestamp("2026-05-13").date()],
            "Departamento": ["A"],
            "Estado": ["Feriado"],
            "Tramos trabajados": ["07:30 - 12:00"],
            "Minutos reales": [270],
            "Minutos redondeados": [270],
            "Minutos extra": [0],
            "Horas extra": ["00:00"],
            "Horas totales": ["04:30"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_colores_estado.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): idx + 1 for idx, h in enumerate(headers)}
    zeller_col = header_to_idx["118 Zeller Lucas Ezequiel"]

    wednesday_row = None
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_idx, column=2).value or "") == "Miercoles":
            wednesday_row = row_idx
            break
    assert wednesday_row is not None

    cell = ws.cell(row=wednesday_row, column=zeller_col)
    rgb = str(getattr(getattr(cell.fill, "fgColor", None), "rgb", "") or "").upper()
    assert rgb.endswith("A9DF8F")


def test_hs_riorda_keeps_exception_status_color_even_if_below_target(tmp_path: Path) -> None:
    daily_df = pd.DataFrame(
        {
            "ID de persona": ["118", "119"],
            "Nombre": ["Zeller Lucas Ezequiel", "Lencina Rocio Pilar"],
            "Fecha": [pd.Timestamp("2026-05-13").date(), pd.Timestamp("2026-05-13").date()],
            "Departamento": ["A", "A"],
            "Estado": ["Vacaciones", "Accidente de trabajo"],
            "Tramos trabajados": ["", ""],
            "Minutos reales": [0, 0],
            "Minutos redondeados": [0, 0],
            "Minutos extra": [0, 0],
            "Horas extra": ["00:00", "00:00"],
            "Horas totales": ["00:00", "00:00"],
        }
    )
    monthly_df = pd.DataFrame(
        columns=[
            "ID de persona",
            "Nombre",
            "Dias trabajados",
            "Minutos totales",
            "Minutos extra",
            "Horas extra",
            "Horas totales",
        ]
    )
    inconsistencies_df = pd.DataFrame(
        columns=["ID de persona", "Nombre", "Fecha", "Tipo de inconsistencia", "Detalle"]
    )

    output = tmp_path / "reporte_hs_riorda_colores_excepciones.xlsx"
    export_report(output, daily_df, monthly_df, inconsistencies_df)

    wb = load_workbook(output)
    ws = wb["hs-riorda"]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): idx + 1 for idx, h in enumerate(headers)}
    zeller_col = header_to_idx["118 Zeller Lucas Ezequiel"]
    rocio_col = header_to_idx["119 Lencina Rocio Pilar"]

    wednesday_row = None
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_idx, column=2).value or "") == "Miercoles":
            wednesday_row = row_idx
            break
    assert wednesday_row is not None

    zeller_rgb = str(
        getattr(getattr(ws.cell(row=wednesday_row, column=zeller_col).fill, "fgColor", None), "rgb", "") or ""
    ).upper()
    rocio_rgb = str(
        getattr(getattr(ws.cell(row=wednesday_row, column=rocio_col).fill, "fgColor", None), "rgb", "") or ""
    ).upper()

    # Vacaciones = amarillo, ART = violeta; no deben volverse naranja.
    assert zeller_rgb.endswith("FFF59D")
    assert rocio_rgb.endswith("8E24AA")
