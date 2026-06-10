from __future__ import annotations

from pathlib import Path
import unicodedata

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from services.schedule_info import ScheduleInfoError, load_schedule_profiles


class ExportError(Exception):
    pass


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill("solid", fgColor="F5F8FC")
STATUS_STYLES = {
    # Base
    "normal": {"fill": None, "font": None},
    "presente": {"fill": None, "font": None},
    "domingo": {"fill": PatternFill("solid", fgColor="D9D9D9"), "font": None},  # gris
    "tarde": {"fill": PatternFill("solid", fgColor="2E7D32"), "font": Font(color="FFFFFF", bold=True)},  # verde oscuro
    "tardanza": {"fill": PatternFill("solid", fgColor="2E7D32"), "font": Font(color="FFFFFF", bold=True)},
    "ausente": {"fill": PatternFill("solid", fgColor="EF9A9A"), "font": Font(color="7F1D1D", bold=True)},  # rojo
    "ausencia": {"fill": PatternFill("solid", fgColor="EF9A9A"), "font": Font(color="7F1D1D", bold=True)},
    # Excepciones solicitadas
    "vacaciones": {"fill": PatternFill("solid", fgColor="FFF59D"), "font": None},  # amarillo
    "estudiar": {"fill": PatternFill("solid", fgColor="B3E5FC"), "font": None},  # celeste
    "capacitacion": {"fill": PatternFill("solid", fgColor="F8BBD0"), "font": None},  # rosado
    "capacitaciones": {"fill": PatternFill("solid", fgColor="F8BBD0"), "font": None},  # rosado
    "suspencion": {"fill": PatternFill("solid", fgColor="90CAF9"), "font": None},  # azul
    "suspension": {"fill": PatternFill("solid", fgColor="90CAF9"), "font": None},  # azul
    "sancion sin goce de sueldo": {"fill": PatternFill("solid", fgColor="90CAF9"), "font": None},  # azul
    "no trabajado": {"fill": PatternFill("solid", fgColor="D7CCC8"), "font": None},  # marron claro
    "licencia": {"fill": PatternFill("solid", fgColor="FFCC80"), "font": None},  # naranja
    "enfermedad": {"fill": PatternFill("solid", fgColor="FFCC80"), "font": None},  # naranja
    "feriado": {"fill": PatternFill("solid", fgColor="A9DF8F"), "font": Font(color="14532D", bold=True)},  # verde manzana
    "accidente de trabajo": {
        "fill": PatternFill("solid", fgColor="8E24AA"),  # violeta
        "font": Font(color="FFFFFF", bold=True),
    },
    "art": {"fill": PatternFill("solid", fgColor="8E24AA"), "font": Font(color="FFFFFF", bold=True)},  # alias
}
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
TOTAL_ROW_BORDER = Border(
    left=Side(style="thin", color="0F172A"),
    right=Side(style="thin", color="0F172A"),
    top=Side(style="medium", color="0F172A"),
    bottom=Side(style="medium", color="0F172A"),
)

SHEET_LAYOUTS = {
    "Diario": {
        "widths": {
            "ID de persona": 14,
            "Nombre": 28,
            "Fecha": 12,
            "Departamento": 22,
            "Estado": 12,
            "Tramos trabajados": 64,
            "Minutos reales": 16,
            "Minutos redondeados": 20,
            "Minutos extra": 16,
            "Horas extra": 14,
            "Horas totales": 14,
        },
        "left_cols": {"Nombre", "Departamento", "Tramos trabajados"},
        "wrap_cols": {"Tramos trabajados"},
        "date_cols": {"Fecha"},
        "priority_sort": ["ID de persona", "Fecha", "Nombre"],
    },
    "Mensual": {
        "widths": {
            "ID de persona": 14,
            "Nombre": 28,
            "Dias trabajados": 16,
            "Minutos totales": 16,
            "Minutos extra": 16,
            "Horas extra": 14,
            "Horas totales": 14,
        },
        "left_cols": {"Nombre"},
        "wrap_cols": set(),
        "date_cols": set(),
        "priority_sort": ["ID de persona", "Nombre"],
    },
    "Liquidar": {
        "widths": {
            "Fecha": 12,
            "Dia": 14,
            "Dia #": 8,
        },
        "left_cols": {"Dia"},
        "wrap_cols": set(),
        "date_cols": {"Fecha"},
        "priority_sort": ["Fecha"],
    },
}

INCIDENCIAS_SHEET = "Incidencias"
INCIDENCIAS_SUMMARY_COLUMNS = ["ID de persona", "Nombre", "Total tardanzas", "Total ausencias"]
INCIDENCIAS_DETAIL_COLUMNS = ["ID de persona", "Nombre", "Fecha", "Estado", "Fichada"]
HS_RIORDA_SHEET = "hs-riorda"
HS_RIORDA_TARGET_EMPLOYEES_ORDERED = [
    ("67", "Madera Adrián"),
    ("113", "Pelliza Roque"),
    ("52", "Mario Butto"),
    ("20", "De Carli Gonzalo"),
    ("51", "Quignard Fernando"),
    ("96", "Stieven Emiliano"),
    ("97", "Gomez Pablo"),
    ("101", "Gimenez Joel"),
    ("100", "Bressan Jorge"),
    ("82", "Godoy Denis"),
    ("109", "Acosta Renzo"),
    ("108", "Bruno Daniel"),
    ("105", "Moreno José Benjamin"),
    ("43", "Vanegas Leandro"),
    ("79", "Zarate Olivera Luis"),
    ("111", "Borgognoni Tomás"),
    ("83", "Zurvera Mirna"),
    ("117", "Mansilla Luis Gabriel"),
    ("119", "Lencina Rocio Pilar"),
    ("118", "Zeller Lucas Ezequiel"),
]
HS_RIORDA_TARGET_NAMES = {name for _, name in HS_RIORDA_TARGET_EMPLOYEES_ORDERED}
HS_RIORDA_MAX_NORMAL_HOURS_BY_ID = {
    "67": 9,
    "113": 9,
    "52": 9,
    "20": 9,
    "51": 9,
    "96": 9,
    "97": 9,
    "101": 9,
    "100": 9,
    "82": 9,
    "109": 9,
    "108": 9,
    "105": 9,
    "43": 9,
    "79": 9,
    "111": 9,
    "83": 7,
    "117": 9,
    "119": 8,
    "118": 4.5,
}
HS_RIORDA_ABSENCE_EXEMPT_EMPLOYEE_IDS = {"118"}
HS_RIORDA_FIXED_WORKING_DAYS_BY_ID = {
    "118": {0, 2, 4},  # lunes, miercoles, viernes
}


def _sort_for_report(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    available = [col for col in columns if col in df.columns]
    if not available or df.empty:
        return df
    return df.sort_values(available, kind="stable").reset_index(drop=True)


def _minutes_to_hhmm(total_minutes: int) -> str:
    hours, minutes = divmod(max(0, int(total_minutes)), 60)
    return f"{hours:02d}:{minutes:02d}"


def _normalize_status(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.strip().lower().split())


def _status_style(status_text: object) -> dict[str, object]:
    key = _normalize_status(status_text)
    if key in STATUS_STYLES:
        return STATUS_STYLES[key]
    if key and key not in {"normal"}:
        # Cualquier excepcion no mapeada explicitamente mantiene verde manzana por defecto.
        return STATUS_STYLES["feriado"]
    return STATUS_STYLES["normal"]


def _normalize_person_name(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return " ".join(text.strip().lower().split())


def _load_employee_catalog(base_dir: Path) -> tuple[list[tuple[str, str, str]], bool]:
    candidates = [base_dir / "date.xlsx", base_dir / "info.xlsx"]
    result: list[tuple[str, str, str]] = []
    by_id: dict[str, str] = {}
    loaded_from_file = False

    def _normalize(text: object) -> str:
        clean = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
        return " ".join(clean.strip().lower().replace("_", " ").split())

    def _clean_id(value: object) -> str:
        if pd.isna(value):
            return ""
        text = str(value).strip()
        if text.lower() in {"", "nan", "none", "nat", "-"}:
            return ""
        if text.endswith(".0") and text[:-2].isdigit():
            return text[:-2]
        return text

    for path in candidates:
        if not path.exists():
            continue
        try:
            with pd.ExcelFile(path) as excel:
                sheet_map = {_normalize(name): name for name in excel.sheet_names}
            sheet_name = sheet_map.get(_normalize("Empleados"))
            if sheet_name is not None:
                df = pd.read_excel(path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(path)
        except Exception:
            continue

        if df.empty:
            continue
        columns = [str(c) for c in df.columns]
        col_map = {_normalize(c): c for c in columns}
        id_col = None
        name_col = None
        for key, original in col_map.items():
            if id_col is None and ("id" == key or "legajo" in key or "id de persona" in key):
                id_col = original
            if name_col is None and ("nombre" == key or "empleado" in key or "name" == key):
                name_col = original
        if id_col is None:
            continue

        for _, row in df.iterrows():
            emp_id = _clean_id(row.get(id_col, ""))
            if not emp_id:
                continue
            emp_name = str(row.get(name_col, "")).strip() if name_col else ""
            if emp_name.lower() in {"", "nan", "none", "nat", "-"}:
                emp_name = f"ID {emp_id}"
            if emp_id not in by_id:
                by_id[emp_id] = emp_name
        loaded_from_file = True
        break

    for emp_id, emp_name in by_id.items():
        result.append((emp_id, emp_name, f"{emp_id} {emp_name}"))
    result.sort(key=lambda item: (item[1].lower(), item[0]))
    return result, loaded_from_file


def _load_working_weekdays_catalog(base_dir: Path) -> dict[str, set[int]]:
    candidates = [base_dir / "date.xlsx", base_dir / "info.xlsx"]
    for path in candidates:
        if not path.exists():
            continue
        try:
            profiles = load_schedule_profiles(path)
        except (ScheduleInfoError, Exception):
            continue
        if not profiles:
            continue
        return {
            str(employee_id).strip(): set(values.get("working_weekdays", {0, 1, 2, 3, 4}))
            for employee_id, values in profiles.items()
        }
    return {}


def _build_tardanzas_sheet(daily_df: pd.DataFrame) -> pd.DataFrame:
    """
    Construye la hoja Tardanzas: una fila por cada día con Estado=Tarde,
    mostrando hora esperada, hora real y minutos de retraso.
    """
    cols = ["ID de persona", "Nombre", "Fecha", "Departamento", "Hora esperada", "Hora real", "Min tarde"]
    if daily_df.empty:
        return pd.DataFrame(columns=cols)

    needed = {"ID de persona", "Nombre", "Fecha", "Estado", "Tramos trabajados"}
    if not needed.issubset(set(daily_df.columns)):
        return pd.DataFrame(columns=cols)

    working = daily_df.copy()
    working["_estado_norm"] = working["Estado"].map(_normalize_status)
    tardy = working[working["_estado_norm"].isin({"tarde", "tardanza"})].copy()
    if tardy.empty:
        return pd.DataFrame(columns=cols)

    # Parsear tramos: "07:35-12:00 [07:30-12:00]" → real=07:35, esperado=07:30
    import re as _re
    def _parse_tramo(tramos_text: str) -> tuple[str, str, int]:
        text = str(tramos_text or "").strip()
        # Primer tramo del día (puede haber varios separados por |||)
        first = text.split("|||")[0].strip()
        # Hora real de entrada: primer HH:MM antes del guión
        real_match = _re.match(r"(\d{2}:\d{2})", first)
        real_start = real_match.group(1) if real_match else ""
        # Hora esperada: primer HH:MM dentro de corchetes
        exp_match = _re.search(r"\[(\d{2}:\d{2})", first)
        exp_start = exp_match.group(1) if exp_match else ""
        # Minutos tarde
        if real_start and exp_start:
            try:
                rh, rm = map(int, real_start.split(":"))
                eh, em = map(int, exp_start.split(":"))
                diff = (rh * 60 + rm) - (eh * 60 + em)
                return exp_start, real_start, max(0, diff)
            except ValueError:
                pass
        return exp_start, real_start, 0

    parsed = tardy["Tramos trabajados"].apply(_parse_tramo)
    tardy["Hora esperada"] = [p[0] for p in parsed]
    tardy["Hora real"] = [p[1] for p in parsed]
    tardy["Min tarde"] = [p[2] for p in parsed]

    result = tardy[["ID de persona", "Nombre", "Fecha", "Departamento",
                     "Hora esperada", "Hora real", "Min tarde"]].copy()
    result = result.sort_values(["ID de persona", "Fecha"], kind="stable").reset_index(drop=True)
    return result


def _build_resumen_sheet(daily_df: pd.DataFrame, monthly_df: pd.DataFrame) -> pd.DataFrame:
    """
    Construye una tabla de dos columnas (Concepto / Valor) con los KPIs del mes.
    """
    rows: list[dict] = []

    def _add(concepto: str, valor: object) -> None:
        rows.append({"Concepto": concepto, "Valor": valor})

    def _sep(titulo: str = "") -> None:
        rows.append({"Concepto": titulo, "Valor": None})

    if daily_df.empty:
        _add("Sin datos", "")
        return pd.DataFrame(rows)

    working = daily_df.copy()
    working["_estado_norm"] = working["Estado"].map(_normalize_status)
    working["_fecha_dt"] = pd.to_datetime(working["Fecha"], errors="coerce")

    # --- Período ---
    fechas = working["_fecha_dt"].dropna()
    if not fechas.empty:
        mes_label = fechas.min().strftime("%B %Y").capitalize()
        fecha_desde = fechas.min().strftime("%d/%m/%Y")
        fecha_hasta = fechas.max().strftime("%d/%m/%Y")
    else:
        mes_label = ""
        fecha_desde = fecha_hasta = ""

    _sep("── PERÍODO ──")
    _add("Mes", mes_label)
    _add("Desde", fecha_desde)
    _add("Hasta", fecha_hasta)

    # --- Empleados ---
    total_empleados = working["ID de persona"].nunique()
    _sep("")
    _sep("── EMPLEADOS ──")
    _add("Empleados en el reporte", total_empleados)

    # --- Asistencia ---
    attendance_norms = {"normal", "tarde", "tardanza", "domingo"}
    dias_trabajados = int(working[working["_estado_norm"].isin(attendance_norms)].shape[0])
    tardanzas = int(working[working["_estado_norm"].isin({"tarde", "tardanza"})].shape[0])
    ausentes = int(working[working["_estado_norm"].isin({"ausente", "ausencia"})].shape[0])
    _sep("")
    _sep("── ASISTENCIA ──")
    _add("Días trabajados (Normal + Tarde)", dias_trabajados)
    _add("Tardanzas", tardanzas)
    _add("Ausencias", ausentes)

    # --- Horas ---
    if not monthly_df.empty and "Minutos totales" in monthly_df.columns:
        total_min_normales = int(monthly_df["Minutos totales"].sum())
        total_min_extra = int(monthly_df["Minutos extra"].sum()) if "Minutos extra" in monthly_df.columns else 0
    else:
        total_min_normales = int(working["Minutos redondeados"].sum()) if "Minutos redondeados" in working.columns else 0
        total_min_extra = int(working["Minutos extra"].sum()) if "Minutos extra" in working.columns else 0

    _sep("")
    _sep("── HORAS ──")
    _add("Horas normales totales", _minutes_to_hhmm(total_min_normales))
    _add("Horas extra totales", _minutes_to_hhmm(total_min_extra))

    # --- Excepciones ---
    excepciones_mask = ~working["_estado_norm"].isin(attendance_norms | {"ausente", "ausencia", ""})
    excepciones = working[excepciones_mask]["_estado_norm"].value_counts()
    if not excepciones.empty:
        _sep("")
        _sep("── EXCEPCIONES ──")
        for exc_tipo, count in excepciones.items():
            _add(str(exc_tipo).title(), int(count))

    # --- Por departamento ---
    if "Departamento" in working.columns:
        dept_trabajados = (
            working[working["_estado_norm"].isin(attendance_norms)]
            .groupby("Departamento")["_estado_norm"]
            .count()
            .sort_values(ascending=False)
        )
        if not dept_trabajados.empty:
            _sep("")
            _sep("── DÍAS TRABAJADOS POR SECTOR ──")
            for dept, count in dept_trabajados.items():
                if str(dept).strip():
                    _add(str(dept), int(count))

    return pd.DataFrame(rows)


def _apply_tardanzas_format(worksheet) -> None:
    widths = [14, 28, 12, 22, 16, 16, 14]
    for col_idx, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    late_fill = STATUS_STYLES["tarde"]["fill"]
    late_font = STATUS_STYLES["tarde"]["font"]

    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, 8):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = late_fill if late_fill else (ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill())
            cell.font = late_font if late_font else Font()
            cell.border = THIN_BORDER
            horizontal = "left" if col_idx in {2, 4} else "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center")
            if col_idx == 3 and cell.value not in ("", None):
                cell.number_format = "DD/MM/YYYY"


def _apply_resumen_format(worksheet) -> None:
    worksheet.column_dimensions["A"].width = 36
    worksheet.column_dimensions["B"].width = 20

    section_fill = PatternFill("solid", fgColor="1F4E78")
    section_font = Font(color="FFFFFF", bold=True)
    value_fill_even = ALT_ROW_FILL

    for row_idx in range(1, worksheet.max_row + 1):
        concepto_cell = worksheet.cell(row=row_idx, column=1)
        value_cell = worksheet.cell(row=row_idx, column=2)
        concepto = str(concepto_cell.value or "")

        if concepto.startswith("──"):
            # Fila de sección
            concepto_cell.fill = section_fill
            concepto_cell.font = section_font
            value_cell.fill = section_fill
            concepto_cell.alignment = Alignment(horizontal="left", vertical="center")
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
            concepto_cell.border = THIN_BORDER
            value_cell.border = THIN_BORDER
        elif concepto == "":
            # Fila vacía separadora
            pass
        else:
            fill = value_fill_even if row_idx % 2 == 0 else PatternFill()
            concepto_cell.fill = fill
            value_cell.fill = fill
            concepto_cell.alignment = Alignment(horizontal="left", vertical="center")
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            concepto_cell.border = THIN_BORDER
            value_cell.border = THIN_BORDER


def _build_incidencias_tables(
    daily_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary_empty = pd.DataFrame(columns=INCIDENCIAS_SUMMARY_COLUMNS)
    detail_empty = pd.DataFrame(columns=INCIDENCIAS_DETAIL_COLUMNS)
    if daily_df.empty:
        return summary_empty, detail_empty

    needed = {"ID de persona", "Nombre", "Fecha", "Estado", "Tramos trabajados"}
    if not needed.issubset(set(daily_df.columns)):
        return summary_empty, detail_empty

    working = daily_df.copy()
    working["ID de persona"] = working["ID de persona"].astype(str).str.strip()
    working["Nombre"] = working["Nombre"].astype(str).str.strip()
    working["_FechaOrden"] = pd.to_datetime(working["Fecha"], errors="coerce")
    working["_estado_norm"] = working["Estado"].map(_normalize_status)
    working["_es_tarde"] = working["_estado_norm"].isin({"tarde", "tardanza"})
    working["_es_ausente"] = working["_estado_norm"].isin({"ausente", "ausencia"})

    employees = (
        working[["ID de persona", "Nombre"]]
        .drop_duplicates()
        .sort_values(["Nombre", "ID de persona"], kind="stable")
        .reset_index(drop=True)
    )
    tardy_counts = (
        working.groupby(["ID de persona", "Nombre"], as_index=False)["_es_tarde"]
        .sum()
        .rename(columns={"_es_tarde": "Total tardanzas"})
    )
    absent_counts = (
        working.groupby(["ID de persona", "Nombre"], as_index=False)["_es_ausente"]
        .sum()
        .rename(columns={"_es_ausente": "Total ausencias"})
    )

    summary = employees.merge(tardy_counts, on=["ID de persona", "Nombre"], how="left").merge(
        absent_counts,
        on=["ID de persona", "Nombre"],
        how="left",
    )
    summary["Total tardanzas"] = summary["Total tardanzas"].fillna(0).astype(int)
    summary["Total ausencias"] = summary["Total ausencias"].fillna(0).astype(int)
    summary = summary[INCIDENCIAS_SUMMARY_COLUMNS]

    detail = working[(working["_es_tarde"]) | (working["_es_ausente"])].copy()
    if detail.empty:
        return summary, detail_empty

    detail = detail.sort_values(
        ["ID de persona", "_FechaOrden", "Nombre"],
        kind="stable",
    )
    detail_df = pd.DataFrame(
        {
            "ID de persona": detail["ID de persona"],
            "Nombre": detail["Nombre"],
            "Fecha": detail["Fecha"],
            "Estado": detail["Estado"].astype(str),
            "Fichada": detail["Tramos trabajados"].fillna("").astype(str),
        }
    )
    return summary, detail_df[INCIDENCIAS_DETAIL_COLUMNS]


def _build_liquidar_sheet(
    daily_df: pd.DataFrame,
    base_dir: Path,
    allowed_employee_names: set[str] | None = None,
    forced_employee_names: list[str] | None = None,
    forced_employees: list[tuple[str, str]] | None = None,
    include_saturdays_with_attendance: bool = True,
    include_extra_summary_rows: bool = True,
    include_total_extra_row: bool = True,
    max_normal_hours_by_employee_id: dict[str, float] | None = None,
    ignore_tardiness_for_normal_hours: bool = False,
    include_weekly_common_rows: bool = False,
    include_weekly_extra_rows: bool = False,
    fill_missing_scheduled_as_absent: bool = False,
    weekly_common_cap_minutes: int | None = None,
    include_blank_row_after_quincena: bool = False,
) -> tuple[pd.DataFrame, dict[tuple[int, int], str]]:
    cols_base = ["Fecha", "Dia", "Dia #"]
    if daily_df.empty:
        return pd.DataFrame(columns=cols_base), {}

    needed = {"Fecha", "ID de persona", "Nombre", "Minutos redondeados", "Minutos extra", "Estado"}
    if not needed.issubset(set(daily_df.columns)):
        return pd.DataFrame(columns=cols_base), {}

    working = daily_df.copy()
    working["_fecha"] = pd.to_datetime(working["Fecha"], errors="coerce")
    working = working[working["_fecha"].notna()].copy()
    if working.empty:
        return pd.DataFrame(columns=cols_base), {}

    def _clean_id(value: object) -> str:
        text = str(value).strip()
        if text.endswith(".0") and text[:-2].isdigit():
            return text[:-2]
        return text

    catalog, loaded_from_file = _load_employee_catalog(base_dir)
    from_daily: dict[str, str] = {}
    for _, row in working.iterrows():
        emp_id = _clean_id(row["ID de persona"])
        emp_name = str(row["Nombre"]).strip()
        if emp_id and emp_id not in from_daily:
            from_daily[emp_id] = emp_name

    catalog_by_id = {emp_id: (emp_name, label) for emp_id, emp_name, label in catalog}
    # Si no hay plantilla de empleados disponible, usa fallback desde el Diario.
    # Si hay date.xlsx/info.xlsx cargado, respeta solo esos empleados.
    if not loaded_from_file:
        for emp_id, emp_name in from_daily.items():
            if emp_id not in catalog_by_id:
                catalog_by_id[emp_id] = (emp_name, f"{emp_id} {emp_name}")

    employees = [(emp_id, name, label) for emp_id, (name, label) in catalog_by_id.items()]
    if forced_employees:
        forced_employees_rows: list[tuple[str, str, str]] = []
        for employee_id, employee_name in forced_employees:
            clean_id = _clean_id(employee_id)
            clean_name = str(employee_name or "").strip()
            label = f"{clean_id} {clean_name}".strip() if clean_id else clean_name
            forced_employees_rows.append((clean_id, clean_name, label))
        employees = forced_employees_rows
    elif forced_employee_names:
        name_to_id: dict[str, str] = {}
        for emp_id, (emp_name, _) in catalog_by_id.items():
            normalized = _normalize_person_name(emp_name)
            if normalized and normalized not in name_to_id:
                name_to_id[normalized] = emp_id
        for emp_id, emp_name in from_daily.items():
            normalized = _normalize_person_name(emp_name)
            if normalized and normalized not in name_to_id:
                name_to_id[normalized] = emp_id

        forced_employees: list[tuple[str, str, str]] = []
        for employee_name in forced_employee_names:
            normalized = _normalize_person_name(employee_name)
            employee_id = name_to_id.get(normalized, "")
            label = f"{employee_id} {employee_name}".strip() if employee_id else employee_name
            forced_employees.append((employee_id, employee_name, label))
        employees = forced_employees
    else:
        if allowed_employee_names:
            allowed_norm = {_normalize_person_name(name) for name in allowed_employee_names}
            employees = [
                item for item in employees
                if _normalize_person_name(item[1]) in allowed_norm
            ]
        employees.sort(key=lambda item: (item[1].lower(), item[0]))
    employee_labels = [item[2] for item in employees]
    working_days_by_employee = (
        _load_working_weekdays_catalog(base_dir)
        if fill_missing_scheduled_as_absent
        else {}
    )

    day_names = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
    by_key: dict[tuple[pd.Timestamp, str], dict] = {}
    by_name_key: dict[tuple[pd.Timestamp, str], dict] = {}
    for _, row in working.iterrows():
        day = row["_fecha"].normalize()
        emp_id = _clean_id(row["ID de persona"])
        emp_name = str(row.get("Nombre", "")).strip()
        by_key[(day, emp_id)] = row
        normalized_name = _normalize_person_name(emp_name)
        if normalized_name:
            by_name_key[(day, normalized_name)] = row

    min_day = working["_fecha"].min().normalize()
    max_day = working["_fecha"].max().normalize()
    attendance_statuses = {"normal", "tarde", "tardanza"}
    saturday_rows = working.loc[working["_fecha"].dt.weekday == 5].copy()
    saturday_rows["_is_attendance"] = False
    if not saturday_rows.empty:
        saturday_rows["_estado_norm"] = saturday_rows["Estado"].map(_normalize_status)
        saturday_rows["_has_punch"] = (
            saturday_rows["Tramos trabajados"].fillna("").astype(str).str.strip() != ""
        )
        saturday_rows["_has_minutes"] = pd.to_numeric(
            saturday_rows["Minutos redondeados"], errors="coerce"
        ).fillna(0).astype(int) > 0
        saturday_rows["_is_attendance"] = (
            saturday_rows["_estado_norm"].isin(attendance_statuses)
            & (saturday_rows["_has_punch"] | saturday_rows["_has_minutes"])
        )
    worked_saturdays = set()
    if include_saturdays_with_attendance:
        worked_saturdays = {
            day.normalize()
            for day in saturday_rows.loc[saturday_rows["_is_attendance"], "_fecha"]
        }
    all_days = [
        day
        for day in pd.date_range(min_day, max_day, freq="D")
        if int(day.weekday()) < 5 or day.normalize() in worked_saturdays
    ]

    rows: list[dict[str, object]] = []
    cell_status_map: dict[tuple[int, int], str] = {}
    totals_minutes_by_employee: dict[str, dict[str, int]] = {
        label: {
            "q1_common": 0,
            "q1_extra": 0,
            "q2_common": 0,
            "q2_extra": 0,
        }
        for _, _, label in employees
    }
    excel_row = 2  # header is row 1

    def _append_day_row(day: pd.Timestamp) -> tuple[dict[str, int], dict[str, int]]:
        nonlocal excel_row
        row_data: dict[str, object] = {
            "Fecha": day.date(),
            "Dia": day_names[int(day.weekday())],
            "Dia #": int(day.day),
        }
        day_common_minutes: dict[str, int] = {label: 0 for _, _, label in employees}
        day_extra_minutes: dict[str, int] = {label: 0 for _, _, label in employees}
        is_saturday = int(day.weekday()) == 5
        attendance_statuses = {"normal", "tarde", "tardanza", "domingo"}
        for emp_index, (emp_id, employee_name, label) in enumerate(employees, start=0):
            weekday = int(day.weekday())
            employee_working_days = working_days_by_employee.get(emp_id)
            if emp_id in HS_RIORDA_FIXED_WORKING_DAYS_BY_ID:
                # Refuerzo operativo: Zeller siempre usa L/Mi/V, incluso si
                # el parseo de date.xlsx cae en un default amplio.
                if employee_working_days is None or employee_working_days == {0, 1, 2, 3, 4}:
                    employee_working_days = set(HS_RIORDA_FIXED_WORKING_DAYS_BY_ID[emp_id])
            if employee_working_days is None:
                employee_working_days = {0, 1, 2, 3, 4}

            if (
                fill_missing_scheduled_as_absent
                and weekday < 5
                and emp_id in HS_RIORDA_ABSENCE_EXEMPT_EMPLOYEE_IDS
                and weekday not in employee_working_days
            ):
                # Zeller en martes/jueves: no mostrar registro ni ausencia.
                row_data[label] = ""
                continue

            rec = by_key.get((day, emp_id))
            if rec is None:
                rec = by_name_key.get((day, _normalize_person_name(employee_name)))
            if rec is None:
                should_mark_absent = False
                if fill_missing_scheduled_as_absent and weekday < 5:
                    if emp_id in HS_RIORDA_ABSENCE_EXEMPT_EMPLOYEE_IDS:
                        # Excepcion operativa: Zeller solo ausente en sus dias laborables.
                        should_mark_absent = weekday in employee_working_days
                    else:
                        # Resto del personal: sin registro en dia habil sale Ausente 0.00.
                        should_mark_absent = True
                if should_mark_absent:
                    row_data[label] = 0.0
                    cell_status_map[(excel_row, 4 + emp_index)] = "Ausente"
                else:
                    row_data[label] = ""
                continue
            status = str(rec.get("Estado", "")).strip()
            status_norm = _normalize_status(status)
            if is_saturday and status_norm and status_norm not in attendance_statuses:
                # Sabado con excepcion cargada: no liquidar horas para ese empleado ese dia.
                row_data[label] = ""
                continue
            total_minutes = int(pd.to_numeric(rec.get("Minutos redondeados", 0), errors="coerce") or 0)
            extra_minutes = int(pd.to_numeric(rec.get("Minutos extra", 0), errors="coerce") or 0)
            normal_minutes = max(0, total_minutes - extra_minutes)
            cap_minutes = None
            if max_normal_hours_by_employee_id:
                max_hours = max_normal_hours_by_employee_id.get(emp_id)
                if max_hours is not None:
                    cap_minutes = int(round(float(max_hours) * 60))
                    # hs-riorda: considerar normales las horas reales hasta el tope diario,
                    # sin descontar "extra" del sistema base para estos perfiles.
                    normal_minutes = min(total_minutes, cap_minutes)
            if ignore_tardiness_for_normal_hours and status_norm in {"tarde", "tardanza"} and cap_minutes is not None:
                # hs-riorda: no descontar por tardanza, usar jornada normal topeada por empleado.
                normal_minutes = cap_minutes
            display_minutes = total_minutes if is_saturday else normal_minutes
            row_data[label] = round(display_minutes / 60, 2)
            day_common_minutes[label] = normal_minutes
            day_extra_minutes[label] = max(0, extra_minutes)
            half = "q1" if int(day.day) <= 15 else "q2"
            totals_minutes_by_employee[label][f"{half}_common"] += normal_minutes
            totals_minutes_by_employee[label][f"{half}_extra"] += max(0, extra_minutes)
            # Employee cols start after A,B,C
            cell_status_map[(excel_row, 4 + emp_index)] = status
        rows.append(row_data)
        excel_row += 1
        return day_common_minutes, day_extra_minutes

    def _hours(minutes: int) -> float:
        return round(max(0, int(minutes)) / 60, 2)

    def _summary_row(label_text: str, metric_key: str) -> dict[str, object]:
        row_data: dict[str, object] = {"Fecha": "", "Dia": label_text, "Dia #": ""}
        for _, _, emp_label in employees:
            row_data[emp_label] = _hours(totals_minutes_by_employee[emp_label][metric_key])
        return row_data

    def _blank_row() -> dict[str, object]:
        row_data: dict[str, object] = {"Fecha": "", "Dia": "", "Dia #": ""}
        for _, _, emp_label in employees:
            row_data[emp_label] = ""
        return row_data

    def _append_week_row(week_index: int, minutes_by_employee: dict[str, int], metric: str) -> None:
        nonlocal excel_row
        week_row: dict[str, object] = {"Fecha": "", "Dia": f"Total Semana {week_index} - {metric}", "Dia #": ""}
        for _, _, emp_label in employees:
            week_row[emp_label] = _hours(minutes_by_employee[emp_label])
        rows.append(week_row)
        excel_row += 1

    def _finalize_week(
        week_index: int,
        week_common_minutes: dict[str, int],
        week_extra_minutes: dict[str, int],
        half_key: str,
    ) -> tuple[int, dict[str, int], dict[str, int]]:
        display_week_common = dict(week_common_minutes)
        display_week_extra = dict(week_extra_minutes)
        cap = int(weekly_common_cap_minutes) if weekly_common_cap_minutes is not None else None
        for label, common_minutes in week_common_minutes.items():
            if cap is not None and common_minutes > cap:
                overflow = common_minutes - cap
                totals_minutes_by_employee[label][f"{half_key}_common"] -= overflow
                totals_minutes_by_employee[label][f"{half_key}_extra"] += overflow
                display_week_common[label] = cap
                display_week_extra[label] = display_week_extra[label] + overflow
        if include_weekly_common_rows:
            _append_week_row(week_index, display_week_common, "Horas Comunes")
        if include_weekly_extra_rows:
            _append_week_row(week_index, display_week_extra, "Horas Extras")
        if include_weekly_common_rows or include_weekly_extra_rows:
            week_index += 1
        return (
            week_index,
            {label: 0 for _, _, label in employees},
            {label: 0 for _, _, label in employees},
        )

    q1_days = [day for day in all_days if int(day.day) <= 15]
    q2_days = [day for day in all_days if int(day.day) > 15]
    week_index = 1
    current_week_common_minutes = {label: 0 for _, _, label in employees}
    current_week_extra_minutes = {label: 0 for _, _, label in employees}

    for day in q1_days:
        day_common, day_extra = _append_day_row(day)
        for label, value in day_common.items():
            current_week_common_minutes[label] += value
        for label, value in day_extra.items():
            current_week_extra_minutes[label] += value
        if int(day.weekday()) == 4:
            (
                week_index,
                current_week_common_minutes,
                current_week_extra_minutes,
            ) = _finalize_week(week_index, current_week_common_minutes, current_week_extra_minutes, "q1")

    if q1_days and (any(current_week_common_minutes.values()) or any(current_week_extra_minutes.values())):
        (
            week_index,
            current_week_common_minutes,
            current_week_extra_minutes,
        ) = _finalize_week(week_index, current_week_common_minutes, current_week_extra_minutes, "q1")

    if q1_days:
        rows.append(_summary_row("1ra Quincena - Horas Comunes", "q1_common"))
        excel_row += 1
        if include_extra_summary_rows:
            rows.append(_summary_row("1ra Quincena - Horas Extras", "q1_extra"))
            excel_row += 1
        if include_blank_row_after_quincena:
            rows.append(_blank_row())
            excel_row += 1

    for day in q2_days:
        day_common, day_extra = _append_day_row(day)
        for label, value in day_common.items():
            current_week_common_minutes[label] += value
        for label, value in day_extra.items():
            current_week_extra_minutes[label] += value
        if int(day.weekday()) == 4:
            (
                week_index,
                current_week_common_minutes,
                current_week_extra_minutes,
            ) = _finalize_week(week_index, current_week_common_minutes, current_week_extra_minutes, "q2")

    if q2_days and (any(current_week_common_minutes.values()) or any(current_week_extra_minutes.values())):
        (
            week_index,
            current_week_common_minutes,
            current_week_extra_minutes,
        ) = _finalize_week(week_index, current_week_common_minutes, current_week_extra_minutes, "q2")

    if q2_days:
        rows.append(_summary_row("2da Quincena - Horas Comunes", "q2_common"))
        excel_row += 1
        if include_extra_summary_rows:
            rows.append(_summary_row("2da Quincena - Horas Extras", "q2_extra"))
            excel_row += 1
        if include_blank_row_after_quincena:
            rows.append(_blank_row())
            excel_row += 1

    total_common_row: dict[str, object] = {"Fecha": "", "Dia": "Total Mes - Horas Comunes", "Dia #": ""}
    for _, _, emp_label in employees:
        total_common_row[emp_label] = _hours(
            totals_minutes_by_employee[emp_label]["q1_common"] + totals_minutes_by_employee[emp_label]["q2_common"]
        )
    rows.append(total_common_row)
    if include_total_extra_row:
        total_extra_row: dict[str, object] = {"Fecha": "", "Dia": "Total Mes - Horas Extras", "Dia #": ""}
        for _, _, emp_label in employees:
            total_extra_row[emp_label] = _hours(
                totals_minutes_by_employee[emp_label]["q1_extra"] + totals_minutes_by_employee[emp_label]["q2_extra"]
            )
        rows.append(total_extra_row)

    liquidar_df = pd.DataFrame(rows, columns=cols_base + employee_labels)
    return liquidar_df, cell_status_map


def _apply_liquidar_format(
    worksheet,
    status_cells: dict[tuple[int, int], str],
    avoid_dark_green: bool = False,
    highlight_below_daily_target: bool = False,
    daily_target_hours_by_employee_id: dict[str, float] | None = None,
) -> None:
    headers = [cell.value for cell in worksheet[1]]
    header_to_idx = {str(value): idx + 1 for idx, value in enumerate(headers) if value is not None}

    worksheet.freeze_panes = "D2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for header, width in SHEET_LAYOUTS["Liquidar"]["widths"].items():
        col_idx = header_to_idx.get(header)
        if col_idx:
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx in range(4, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 17

    for row_idx in range(2, worksheet.max_row + 1):
        first_col_value = worksheet.cell(row=row_idx, column=1).value
        is_daily_row = pd.notna(pd.to_datetime(first_col_value, errors="coerce"))
        row_label = str(worksheet.cell(row=row_idx, column=2).value or "").strip().lower()
        is_total_separator_row = (not is_daily_row) and (
            "total semana" in row_label
            or "quincena" in row_label
            or "total mes" in row_label
        )
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = TOTAL_ROW_BORDER if is_total_separator_row else THIN_BORDER
            if is_daily_row and row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            if (not is_daily_row) and row_label:
                if "quincena" in row_label:
                    cell.fill = PatternFill("solid", fgColor="E9D5FF")
                elif "horas extras" in row_label:
                    cell.fill = PatternFill("solid", fgColor="FFD7B5")
                elif "horas comunes" in row_label:
                    cell.fill = PatternFill("solid", fgColor="C7E0FF")
                # Quincenas y total mensual: resaltar letras y valores.
                cell.font = Font(color="0F172A", bold=True)

            if col_idx == 1 and is_daily_row and cell.value not in ("", None):
                cell.number_format = "DD/MM/YYYY"
            if col_idx >= 4:
                cell.number_format = "0.00"

            horizontal = "center"
            if col_idx in {2}:
                horizontal = "left"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center")

    # Color cells by status for liquidation view
    for (row_idx, col_idx), status in status_cells.items():
        status_norm = _normalize_status(status)
        if avoid_dark_green and status_norm in {"tarde", "tardanza"}:
            continue
        style = _status_style(status)
        fill = style.get("fill")
        font = style.get("font")
        cell = worksheet.cell(row=row_idx, column=col_idx)
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font

    if highlight_below_daily_target and daily_target_hours_by_employee_id:
        highlight_fill = PatternFill("solid", fgColor="FFB74D")
        orange_allowed_statuses = {"normal", "presente", "tarde", "tardanza", "domingo"}
        for row_idx in range(2, worksheet.max_row + 1):
            first_col_value = worksheet.cell(row=row_idx, column=1).value
            is_daily_row = pd.notna(pd.to_datetime(first_col_value, errors="coerce"))
            if not is_daily_row:
                continue
            for col_idx in range(4, worksheet.max_column + 1):
                header = str(worksheet.cell(row=1, column=col_idx).value or "").strip()
                employee_id = header.split(" ", 1)[0].strip() if header else ""
                if not employee_id:
                    continue
                target_hours = daily_target_hours_by_employee_id.get(employee_id)
                if target_hours is None:
                    continue
                status_text = status_cells.get((row_idx, col_idx), "")
                status_norm = _normalize_status(status_text)
                if status_norm not in orange_allowed_statuses:
                    # Mantener colores de estados especiales (vacaciones, ART, feriado, etc).
                    continue
                cell = worksheet.cell(row=row_idx, column=col_idx)
                raw_value = cell.value
                if raw_value in ("", None):
                    continue
                try:
                    worked_hours = float(raw_value)
                except Exception:
                    continue
                if worked_hours + 1e-9 < float(target_hours):
                    cell.fill = highlight_fill
                    cell.font = Font(color="0F172A", bold=True)

    _append_liquidar_legend(worksheet, avoid_dark_green=avoid_dark_green)


def _append_liquidar_legend(worksheet, avoid_dark_green: bool = False) -> None:
    legend_items = [
        ("Domingo", "domingo"),
        ("Ausente", "ausente"),
        ("Vacaciones", "vacaciones"),
        ("Estudiar", "estudiar"),
        ("Capacitacion", "capacitacion"),
        ("Suspencion", "suspencion"),
        ("No trabajado", "no trabajado"),
        ("Licencia", "licencia"),
        ("Feriado", "feriado"),
        ("Accidente de trabajo", "accidente de trabajo"),
        ("Tardanza", "tardanza"),
    ]

    start_row = worksheet.max_row + 2
    title_cell = worksheet.cell(row=start_row, column=1)
    title_cell.value = "Leyenda de estados"
    title_cell.fill = HEADER_FILL
    title_cell.font = HEADER_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = THIN_BORDER
    worksheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=4,
    )

def _apply_incidencias_format(worksheet, summary_rows: int, detail_header_row: int) -> None:
    headers_row_1 = [cell.value for cell in worksheet[1]]
    header_row_1_idx = {str(value): idx + 1 for idx, value in enumerate(headers_row_1) if value is not None}

    worksheet.freeze_panes = "A2"

    for col_idx, width in enumerate((14, 30, 16, 16, 56), start=1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    # Header resumen (fila 1)
    for cell in worksheet[1]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Filas resumen
    for row_idx in range(2, summary_rows + 2):
        for col_idx in range(1, 5):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            horizontal = "left" if col_idx == 2 else "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center")

    # Header detalle
    for col_idx in range(1, 6):
        cell = worksheet.cell(row=detail_header_row, column=col_idx)
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Detalle filas
    for row_idx in range(detail_header_row + 1, worksheet.max_row + 1):
        status_text = worksheet.cell(row=row_idx, column=4).value
        style = _status_style(status_text)
        row_fill = style.get("fill")
        row_font = style.get("font")
        for col_idx in range(1, 6):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if row_fill is not None:
                cell.fill = row_fill
            elif row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            if row_font is not None:
                cell.font = row_font
            if col_idx == 3 and cell.value not in ("", None):
                cell.number_format = "DD/MM/YYYY"
            if col_idx in {2, 5}:
                horizontal = "left"
            else:
                horizontal = "center"
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical="top",
                wrap_text=(col_idx == 5),
            )

    # Keep filter on summary header row only
    if header_row_1_idx:
        worksheet.auto_filter.ref = f"A1:D{max(1, summary_rows + 1)}"


def _apply_sheet_format(worksheet, sheet_name: str) -> None:
    config = SHEET_LAYOUTS[sheet_name]
    headers = [cell.value for cell in worksheet[1]]
    header_to_idx = {str(value): idx + 1 for idx, value in enumerate(headers) if value is not None}
    status_col_idx = header_to_idx.get("Estado")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for header, width in config["widths"].items():
        col_idx = header_to_idx.get(header)
        if col_idx:
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx in range(2, worksheet.max_row + 1):
        row_fill = None
        row_font = None
        row_status = ""
        if sheet_name == "Diario" and status_col_idx is not None:
            row_status = str(worksheet.cell(row=row_idx, column=status_col_idx).value or "").strip().lower()
            style = _status_style(row_status)
            row_fill = style.get("fill")
            row_font = style.get("font")

        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if row_fill is not None:
                cell.fill = row_fill
            elif row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

        for header, col_idx in header_to_idx.items():
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER

            horizontal = "left" if header in config["left_cols"] else "center"
            wrap = header in config["wrap_cols"]
            cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=wrap)
            if row_font is not None:
                cell.font = row_font

            if header in config["date_cols"] and cell.value not in ("", None):
                cell.number_format = "DD/MM/YYYY"


def export_report(
    output_path: str | Path,
    daily_df: pd.DataFrame,
    monthly_df: pd.DataFrame,
    inconsistencies_df: pd.DataFrame,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output.with_suffix(f"{output.suffix}.tmp")

    diario_export = _sort_for_report(daily_df.copy(), SHEET_LAYOUTS["Diario"]["priority_sort"])
    mensual_export = _sort_for_report(monthly_df.copy(), SHEET_LAYOUTS["Mensual"]["priority_sort"])
    tardanzas_export = _build_tardanzas_sheet(diario_export)
    resumen_export = _build_resumen_sheet(diario_export, mensual_export)
    incidencias_summary, incidencias_detail = _build_incidencias_tables(diario_export)
    liquidar_export, liquidar_status_cells = _build_liquidar_sheet(
        diario_export,
        output.parent,
    )
    hs_riorda_export, hs_riorda_status_cells = _build_liquidar_sheet(
        diario_export,
        output.parent,
        allowed_employee_names=HS_RIORDA_TARGET_NAMES,
        forced_employees=HS_RIORDA_TARGET_EMPLOYEES_ORDERED,
        include_saturdays_with_attendance=False,
        include_extra_summary_rows=True,
        include_total_extra_row=True,
        max_normal_hours_by_employee_id=HS_RIORDA_MAX_NORMAL_HOURS_BY_ID,
        ignore_tardiness_for_normal_hours=True,
        include_weekly_common_rows=True,
        include_weekly_extra_rows=True,
        fill_missing_scheduled_as_absent=True,
        weekly_common_cap_minutes=44 * 60,
        include_blank_row_after_quincena=True,
    )
    try:
        with pd.ExcelWriter(temp_output, engine="openpyxl") as writer:
            diario_export.to_excel(writer, sheet_name="Diario", index=False)
            mensual_export.to_excel(writer, sheet_name="Mensual", index=False)
            resumen_export.to_excel(writer, sheet_name="Resumen", index=False)
            if not tardanzas_export.empty:
                tardanzas_export.to_excel(writer, sheet_name="Tardanzas", index=False)
            incidencias_summary.to_excel(writer, sheet_name=INCIDENCIAS_SHEET, index=False)
            detail_startrow = len(incidencias_summary) + 3
            incidencias_detail.to_excel(
                writer,
                sheet_name=INCIDENCIAS_SHEET,
                index=False,
                startrow=detail_startrow,
            )
            liquidar_export.to_excel(writer, sheet_name="Liquidar", index=False)
            hs_riorda_export.to_excel(writer, sheet_name=HS_RIORDA_SHEET, index=False)

            workbook = writer.book
            for sheet_name in ("Diario", "Mensual"):
                _apply_sheet_format(workbook[sheet_name], sheet_name)
            _apply_resumen_format(workbook["Resumen"])
            if "Tardanzas" in workbook.sheetnames:
                _apply_tardanzas_format(workbook["Tardanzas"])
            _apply_incidencias_format(
                workbook[INCIDENCIAS_SHEET],
                summary_rows=len(incidencias_summary),
                detail_header_row=detail_startrow + 1,
            )
            _apply_liquidar_format(workbook["Liquidar"], liquidar_status_cells)
            _apply_liquidar_format(
                workbook[HS_RIORDA_SHEET],
                hs_riorda_status_cells,
                avoid_dark_green=True,
                highlight_below_daily_target=True,
                daily_target_hours_by_employee_id=HS_RIORDA_MAX_NORMAL_HOURS_BY_ID,
            )

        temp_output.replace(output)
    except Exception as exc:
        if temp_output.exists():
            temp_output.unlink(missing_ok=True)
        raise ExportError(f"No se pudo exportar el Excel: {exc}") from exc

    return output
